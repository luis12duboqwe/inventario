"""Generadores de reportes PDF y Excel para inventario."""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .. import schemas


def _format_currency(value: Decimal | float | int) -> str:
    normalized = float(value)
    return f"${normalized:,.2f}"


def _format_units(value: int) -> str:
    return f"{value:,}".replace(",", ".")


def _build_table(table_data: list[list[str]]) -> Table:
    table = Table(table_data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#e2e8f0")),
                ("LINEABOVE", (0, 0), (-1, 0), 1, colors.HexColor("#38bdf8")),
                ("LINEBELOW", (0, -1), (-1, -1), 1, colors.HexColor("#38bdf8")),
                ("ALIGN", (1, 1), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )
    return table


def _build_document(title: str) -> tuple[list, SimpleDocTemplate, BytesIO]:  # type: ignore[type-arg]
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, title=title)
    styles = getSampleStyleSheet()
    now_label = datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")

    heading_style = ParagraphStyle(
        "Heading1Softmobile",
        parent=styles["Heading1"],
        textColor=colors.HexColor("#38bdf8"),
    )

    elements: list = [  # type: ignore[var-annotated]
        Paragraph("Softmobile 2025 — Reportes de inventario", heading_style),
        Spacer(1, 12),
        Paragraph(f"Generado automáticamente el {now_label}", styles["Normal"]),
        Spacer(1, 18),
    ]

    return elements, document, buffer


def render_inventory_current_pdf(report: schemas.InventoryCurrentReport) -> bytes:
    elements, document, buffer = _build_document("Softmobile - Existencias actuales")
    styles = getSampleStyleSheet()

    elements.append(
        Paragraph(
            "Existencias consolidadas por sucursal y totales corporativos.",
            styles["Italic"],
        )
    )
    elements.append(Spacer(1, 12))

    totals = report.totals
    summary_table = _build_table(
        [
            ["Indicador", "Valor"],
            ["Sucursales evaluadas", str(totals.stores)],
            ["Dispositivos catalogados", str(totals.devices)],
            ["Unidades totales", _format_units(totals.total_units)],
            ["Valor consolidado", _format_currency(totals.total_value)],
        ]
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    table_data: list[list[str]] = [
        ["Sucursal", "Dispositivos", "Unidades", "Valor total"],
    ]
    for store in report.stores:
        table_data.append(
            [
                store.store_name,
                str(store.device_count),
                _format_units(store.total_units),
                _format_currency(store.total_value),
            ]
        )

    elements.append(_build_table(table_data))

    document.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def render_inventory_value_pdf(report: schemas.InventoryValueReport) -> bytes:
    elements, document, buffer = _build_document("Softmobile - Valor de inventario")
    styles = getSampleStyleSheet()
    elements.append(
        Paragraph(
            "Detalle de valor de venta frente a costo y margen proyectado.",
            styles["Italic"],
        )
    )
    elements.append(Spacer(1, 12))

    totals = report.totals
    summary_table = _build_table(
        [
            ["Indicador", "Valor"],
            ["Valor total", _format_currency(totals.valor_total)],
            ["Costo estimado", _format_currency(totals.valor_costo)],
            ["Margen proyectado", _format_currency(totals.margen_total)],
        ]
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    table_data: list[list[str]] = [
        ["Sucursal", "Valor total", "Costo", "Margen"],
    ]
    for store in report.stores:
        table_data.append(
            [
                store.store_name,
                _format_currency(store.valor_total),
                _format_currency(store.valor_costo),
                _format_currency(store.margen_total),
            ]
        )

    elements.append(_build_table(table_data))

    document.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def render_inventory_movements_pdf(report: schemas.InventoryMovementsReport) -> bytes:
    elements, document, buffer = _build_document("Softmobile - Movimientos de inventario")
    styles = getSampleStyleSheet()
    elements.append(
        Paragraph(
            "Entradas, salidas y ajustes consolidados por periodo y tipo de movimiento.",
            styles["Italic"],
        )
    )
    elements.append(Spacer(1, 12))

    resumen = report.resumen
    summary_table = _build_table(
        [
            ["Indicador", "Valor"],
            ["Total de movimientos", _format_units(resumen.total_movimientos)],
            ["Unidades movilizadas", _format_units(resumen.total_unidades)],
            ["Valor total movilizado", _format_currency(resumen.total_valor)],
        ]
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    type_table: list[list[str]] = [
        ["Tipo", "Cantidad", "Valor"],
    ]
    for entry in resumen.por_tipo:
        type_table.append(
            [
                entry.tipo_movimiento.value.upper(),
                _format_units(entry.total_cantidad),
                _format_currency(entry.total_valor),
            ]
        )
    elements.append(_build_table(type_table))
    elements.append(Spacer(1, 18))

    period_table: list[list[str]] = [
        ["Fecha", "Tipo", "Cantidad", "Valor"],
    ]
    for entry in report.periodos:
        period_table.append(
            [
                entry.periodo.isoformat(),
                entry.tipo_movimiento.value.upper(),
                _format_units(entry.total_cantidad),
                _format_currency(entry.total_valor),
            ]
        )
    if len(period_table) > 1:
        elements.append(_build_table(period_table))
        elements.append(Spacer(1, 18))

    detail_table: list[list[str]] = [
        [
            "ID",
            "Fecha",
            "Tipo",
            "Cantidad",
            "Valor",
            "Destino",
            "Origen",
            "Usuario",
            "Comentario",
        ]
    ]
    for movement in report.movimientos:
        detail_table.append(
            [
                str(movement.id),
                movement.fecha.strftime("%Y-%m-%d %H:%M"),
                movement.tipo_movimiento.value.upper(),
                _format_units(movement.cantidad),
                _format_currency(movement.valor_total),
                movement.tienda_destino or "-",
                movement.tienda_origen or "-",
                movement.usuario or "-",
                movement.comentario or "-",
            ]
        )
    elements.append(_build_table(detail_table))

    document.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def render_top_products_pdf(report: schemas.TopProductsReport) -> bytes:
    elements, document, buffer = _build_document("Softmobile - Productos más vendidos")
    styles = getSampleStyleSheet()
    elements.append(
        Paragraph(
            "Desempeño de productos destacados con ingresos y margen estimado.",
            styles["Italic"],
        )
    )
    elements.append(Spacer(1, 12))

    summary_table = _build_table(
        [
            ["Indicador", "Valor"],
            ["Total unidades", _format_units(report.total_unidades)],
            ["Ingresos totales", _format_currency(report.total_ingresos)],
        ]
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    table_data: list[list[str]] = [
        ["SKU", "Producto", "Sucursal", "Unidades", "Ingresos", "Margen"],
    ]
    for item in report.items:
        table_data.append(
            [
                item.sku,
                item.nombre,
                item.store_name,
                _format_units(item.unidades_vendidas),
                _format_currency(item.ingresos_totales),
                _format_currency(item.margen_estimado),
            ]
        )

    elements.append(_build_table(table_data))

    document.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def _apply_header_style(cell) -> None:
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill("solid", fgColor="0f172a")
    cell.alignment = Alignment(horizontal="center")


def _autosize_columns(sheet) -> None:
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max_length + 2


def build_inventory_current_excel(report: schemas.InventoryCurrentReport) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Existencias"

    sheet.append(["Resumen", "Valor"])
    _apply_header_style(sheet["A1"])
    _apply_header_style(sheet["B1"])
    sheet.append(["Sucursales evaluadas", report.totals.stores])
    sheet.append(["Dispositivos catalogados", report.totals.devices])
    sheet.append(["Unidades totales", report.totals.total_units])
    sheet.append(["Valor consolidado", report.totals.total_value])
    sheet.append([])

    sheet.append(["Sucursal", "Dispositivos", "Unidades", "Valor total"])
    for cell in sheet[sheet.max_row]:
        _apply_header_style(cell)
    for store in report.stores:
        sheet.append(
            [
                store.store_name,
                store.device_count,
                store.total_units,
                float(store.total_value),
            ]
        )

    _autosize_columns(sheet)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_inventory_value_excel(report: schemas.InventoryValueReport) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Valor"

    sheet.append(["Indicador", "Valor"])
    _apply_header_style(sheet["A1"])
    _apply_header_style(sheet["B1"])
    sheet.append(["Valor total", float(report.totals.valor_total)])
    sheet.append(["Costo estimado", float(report.totals.valor_costo)])
    sheet.append(["Margen proyectado", float(report.totals.margen_total)])
    sheet.append([])

    sheet.append(["Sucursal", "Valor total", "Costo", "Margen"])
    for cell in sheet[sheet.max_row]:
        _apply_header_style(cell)
    for store in report.stores:
        sheet.append(
            [
                store.store_name,
                float(store.valor_total),
                float(store.valor_costo),
                float(store.margen_total),
            ]
        )

    _autosize_columns(sheet)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_inventory_movements_excel(report: schemas.InventoryMovementsReport) -> BytesIO:
    workbook = Workbook()
    resumen_sheet = workbook.active
    resumen_sheet.title = "Resumen"

    resumen_sheet.append(["Indicador", "Valor"])
    _apply_header_style(resumen_sheet["A1"])
    _apply_header_style(resumen_sheet["B1"])
    resumen_sheet.append(["Total de movimientos", report.resumen.total_movimientos])
    resumen_sheet.append(["Unidades movilizadas", report.resumen.total_unidades])
    resumen_sheet.append(["Valor total movilizado", float(report.resumen.total_valor)])

    resumen_sheet.append([])
    resumen_sheet.append(["Tipo", "Cantidad", "Valor"])
    for cell in resumen_sheet[resumen_sheet.max_row]:
        _apply_header_style(cell)
    for entry in report.resumen.por_tipo:
        resumen_sheet.append(
            [
                entry.tipo_movimiento.value.upper(),
                entry.total_cantidad,
                float(entry.total_valor),
            ]
        )
    _autosize_columns(resumen_sheet)

    period_sheet = workbook.create_sheet("Periodos")
    period_sheet.append(["Fecha", "Tipo", "Cantidad", "Valor"])
    for cell in period_sheet[1]:
        _apply_header_style(cell)
    for entry in report.periodos:
        period_sheet.append(
            [
                entry.periodo.isoformat(),
                entry.tipo_movimiento.value.upper(),
                entry.total_cantidad,
                float(entry.total_valor),
            ]
        )
    _autosize_columns(period_sheet)

    detail_sheet = workbook.create_sheet("Detalle")
    detail_sheet.append(
        [
            "ID",
            "Fecha",
            "Tipo",
            "Cantidad",
            "Valor",
            "Destino",
            "Origen",
            "Usuario",
            "Comentario",
        ]
    )
    for cell in detail_sheet[1]:
        _apply_header_style(cell)
    for movement in report.movimientos:
        detail_sheet.append(
            [
                movement.id,
                movement.fecha.strftime("%Y-%m-%d %H:%M"),
                movement.tipo_movimiento.value.upper(),
                movement.cantidad,
                float(movement.valor_total),
                movement.tienda_destino or "-",
                movement.tienda_origen or "-",
                movement.usuario or "-",
                movement.comentario or "-",
            ]
        )
    _autosize_columns(detail_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_top_products_excel(report: schemas.TopProductsReport) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Top productos"

    sheet.append(["Indicador", "Valor"])
    _apply_header_style(sheet["A1"])
    _apply_header_style(sheet["B1"])
    sheet.append(["Total unidades", report.total_unidades])
    sheet.append(["Ingresos totales", float(report.total_ingresos)])
    sheet.append([])

    sheet.append(["SKU", "Producto", "Sucursal", "Unidades", "Ingresos", "Margen"])
    for cell in sheet[sheet.max_row]:
        _apply_header_style(cell)
    for item in report.items:
        sheet.append(
            [
                item.sku,
                item.nombre,
                item.store_name,
                item.unidades_vendidas,
                float(item.ingresos_totales),
                float(item.margen_estimado),
            ]
        )

    _autosize_columns(sheet)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
