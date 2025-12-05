from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .. import schemas

DARK_BACKGROUND = colors.HexColor("#0f172a")
DARK_SURFACE = colors.HexColor("#111827")
ACCENT = colors.HexColor("#38bdf8")
TEXT_PRIMARY = colors.HexColor("#e2e8f0")


def _format_timestamp(value: datetime | None) -> str:
    if value is None:
        return "—"
    return value.astimezone().strftime("%d/%m/%Y %H:%M")


def build_conflict_report(
    conflicts: Sequence[schemas.SyncConflictLog],
    filters: schemas.SyncConflictReportFilters,
) -> schemas.SyncConflictReport:
    critical = sum(1 for entry in conflicts if entry.severity == schemas.SyncBranchHealth.CRITICAL)
    warning = sum(1 for entry in conflicts if entry.severity == schemas.SyncBranchHealth.WARNING)
    affected_skus = len({entry.sku for entry in conflicts})

    totals = schemas.SyncConflictReportTotals(
        count=len(conflicts),
        critical=critical,
        warning=warning,
        affected_skus=affected_skus,
    )

    return schemas.SyncConflictReport(
        generated_at=datetime.now(timezone.utc),
        filters=filters,
        totals=totals,
        items=list(conflicts),
    )


def render_conflict_report_pdf(report: schemas.SyncConflictReport) -> bytes:
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, title="Conflictos de sincronización")
    styles = getSampleStyleSheet()

    heading_style = ParagraphStyle(
        "HeadingSoftmobileConflicts",
        parent=styles["Heading1"],
        textColor=ACCENT,
    )
    now_label = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")

    elements: list = [  # type: ignore[var-annotated]
        Paragraph("Softmobile 2025 — Conflictos de sincronización", heading_style),
        Spacer(1, 12),
        Paragraph(f"Generado automáticamente el {now_label}", styles["Normal"]),
        Spacer(1, 18),
    ]

    filters_parts: list[str] = []
    if report.filters.store_id:
        filters_parts.append(f"Sucursal filtrada #{report.filters.store_id}")
    if report.filters.severity:
        filters_parts.append(f"Severidad: {report.filters.severity.value}")
    if report.filters.date_from or report.filters.date_to:
        start = (
            report.filters.date_from.strftime("%d/%m/%Y")
            if report.filters.date_from
            else "∞"
        )
        end = (
            report.filters.date_to.strftime("%d/%m/%Y")
            if report.filters.date_to
            else "∞"
        )
        filters_parts.append(f"Periodo: {start} → {end}")

    filters_label = ", ".join(filters_parts) if filters_parts else "Sin filtros aplicados"
    elements.append(Paragraph(filters_label, styles["Italic"]))
    elements.append(Spacer(1, 12))

    summary_table = Table(
        [
            ["Indicador", "Valor"],
            ["Eventos analizados", str(report.totals.count)],
            ["Críticos", str(report.totals.critical)],
            ["Advertencias", str(report.totals.warning)],
            ["SKU afectados", str(report.totals.affected_skus)],
        ]
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), DARK_BACKGROUND),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), DARK_SURFACE),
                ("TEXTCOLOR", (0, 1), (-1, -1), TEXT_PRIMARY),
                ("LINEABOVE", (0, 0), (-1, 0), 1, ACCENT),
                ("LINEBELOW", (0, -1), (-1, -1), 1, ACCENT),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ]
        )
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    detail_header = [
        "SKU",
        "Producto",
        "Severidad",
        "Diferencia",
        "Detectado",
        "Sucursales con máximo",
        "Sucursales con mínimo",
    ]
    detail_rows = [detail_header]
    for entry in report.items:
        max_labels = ", ".join(
            f"{store.store_name} ({store.quantity})" for store in entry.stores_max
        ) or "—"
        min_labels = ", ".join(
            f"{store.store_name} ({store.quantity})" for store in entry.stores_min
        ) or "—"
        detail_rows.append(
            [
                entry.sku,
                entry.product_name or "Sin descripción",
                entry.severity.value,
                str(entry.difference),
                _format_timestamp(entry.detected_at),
                max_labels,
                min_labels,
            ]
        )

    detail_table = Table(detail_rows, hAlign="LEFT")
    detail_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), DARK_BACKGROUND),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), DARK_SURFACE),
                ("TEXTCOLOR", (0, 1), (-1, -1), TEXT_PRIMARY),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#1f2937")),
            ]
        )
    )
    elements.append(detail_table)

    document.build(elements)
    return buffer.getvalue()


def render_conflict_report_excel(report: schemas.SyncConflictReport) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Conflictos"

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="0f172a")

    sheet.append(
        [
            "SKU",
            "Producto",
            "Severidad",
            "Diferencia",
            "Detectado",
            "Sucursales con máximo",
            "Sucursales con mínimo",
        ]
    )
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    for entry in report.items:
        max_labels = ", ".join(
            f"{store.store_name} ({store.quantity})" for store in entry.stores_max
        ) or "—"
        min_labels = ", ".join(
            f"{store.store_name} ({store.quantity})" for store in entry.stores_min
        ) or "—"
        sheet.append(
            [
                entry.sku,
                entry.product_name or "Sin descripción",
                entry.severity.value,
                entry.difference,
                _format_timestamp(entry.detected_at),
                max_labels,
                min_labels,
            ]
        )

    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 24

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
