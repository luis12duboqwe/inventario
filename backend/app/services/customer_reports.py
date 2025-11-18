from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .. import schemas
from .locale_helpers import format_dual_currency


def _format_currency(value: float) -> str:
    return format_dual_currency(value)


def _build_pdf_table(data: list[list[str]]) -> Table:
    table = Table(data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#e2e8f0")),
                ("LINEBEFORE", (0, 0), (0, -1), 1, colors.HexColor("#1e293b")),
                ("LINEAFTER", (-1, 0), (-1, -1), 1, colors.HexColor("#1e293b")),
                ("LINEABOVE", (0, 0), (-1, 0), 1.2, colors.HexColor("#38bdf8")),
                ("LINEBELOW", (0, -1), (-1, -1), 1.2, colors.HexColor("#38bdf8")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return table


def _build_pdf_document(title: str) -> tuple[list, SimpleDocTemplate, BytesIO]:  # type: ignore[type-arg]
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, title=title)
    styles = getSampleStyleSheet()

    heading_style = ParagraphStyle(
        "HeadingSoftmobile",
        parent=styles["Heading1"],
        textColor=colors.HexColor("#38bdf8"),
        fontSize=18,
    )

    now_label = datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")
    elements: list = [  # type: ignore[var-annotated]
        Paragraph("Softmobile 2025 — Reporte de clientes", heading_style),
        Spacer(1, 12),
        Paragraph(f"Generado automáticamente el {now_label}", styles["Normal"]),
        Spacer(1, 18),
    ]
    return elements, document, buffer


def render_customer_portfolio_pdf(report: schemas.CustomerPortfolioReport) -> bytes:
    elements, document, buffer = _build_pdf_document("Softmobile - Reporte de clientes")
    styles = getSampleStyleSheet()

    filters_parts: list[str] = [
        "Portafolio: clientes morosos" if report.category == "delinquent" else "Portafolio: clientes frecuentes"
    ]
    if report.filters.date_from:
        filters_parts.append(f"Desde: {report.filters.date_from.strftime('%d/%m/%Y')}")
    if report.filters.date_to:
        filters_parts.append(f"Hasta: {report.filters.date_to.strftime('%d/%m/%Y')}")
    filters_parts.append(f"Límite: {report.filters.limit}")
    filters_text = " | ".join(filters_parts)

    elements.append(Paragraph(filters_text, styles["Italic"]))
    elements.append(Spacer(1, 12))

    summary_table = _build_pdf_table(
        [
            ["Indicador", "Valor"],
            ["Clientes incluidos", str(report.totals.customers)],
            ["Marcados morosos", str(report.totals.moroso_flagged)],
            ["Deuda total", _format_currency(report.totals.outstanding_debt)],
            ["Ventas acumuladas", _format_currency(report.totals.sales_total)],
        ]
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    detail_table: list[list[str]] = [
        [
            "Cliente",
            "Estado",
            "Tipo",
            "Crédito",
            "Saldo",
            "Disponible",
            "Ventas",
            "Operaciones",
            "Última venta",
        ]
    ]

    for item in report.items:
        last_sale_label = "—"
        if item.last_sale_at:
            last_sale_label = item.last_sale_at.strftime("%d/%m/%Y %H:%M")
        detail_table.append(
            [
                item.name,
                item.status.title(),
                item.customer_type.title(),
                _format_currency(item.credit_limit),
                _format_currency(item.outstanding_debt),
                _format_currency(item.available_credit),
                _format_currency(item.sales_total),
                str(item.sales_count),
                last_sale_label,
            ]
        )

    elements.append(_build_pdf_table(detail_table))
    document.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def render_customer_statement_pdf(report: schemas.CustomerStatementReport) -> bytes:
    elements, document, buffer = _build_pdf_document("Softmobile - Estado de cuenta")
    styles = getSampleStyleSheet()

    customer = report.customer
    summary = report.summary
    header_text = f"Cliente: {customer.name}"
    if customer.customer_type:
        header_text += f" · Tipo {customer.customer_type.title()}"
    elements.append(Paragraph(header_text, styles["Heading2"]))
    elements.append(Spacer(1, 12))

    contact_lines = [
        ["Correo", customer.email or "—"],
        ["Teléfono", customer.phone or "—"],
        ["Saldo pendiente", _format_currency(summary.total_outstanding)],
        ["Crédito disponible", _format_currency(summary.available_credit)],
        ["Crédito autorizado", _format_currency(summary.credit_limit)],
        [
            "Último pago",
            summary.last_payment_at.strftime("%d/%m/%Y %H:%M")
            if summary.last_payment_at
            else "—",
        ],
        [
            "Próximo vencimiento",
            summary.next_due_date.strftime("%d/%m/%Y %H:%M")
            if summary.next_due_date
            else "—",
        ],
        [
            "Días promedio en cartera",
            f"{summary.average_days_outstanding:.1f}",
        ],
    ]
    summary_table = _build_pdf_table([["Dato", "Valor"], *contact_lines])
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    ledger_table: list[list[str]] = [["Fecha", "Descripción", "Referencia", "Monto", "Saldo"]]
    for line in report.lines:
        amount_label = _format_currency(abs(line.amount))
        if line.amount < 0:
            amount_label = f"-{amount_label}"
        balance_label = _format_currency(line.balance_after)
        ledger_table.append(
            [
                line.created_at.strftime("%d/%m/%Y %H:%M"),
                line.description,
                line.reference or "—",
                amount_label,
                balance_label,
            ]
        )

    elements.append(_build_pdf_table(ledger_table))
    elements.append(Spacer(1, 12))
    elements.append(
        Paragraph(
            f"Estado generado el {report.generated_at.strftime('%d/%m/%Y %H:%M UTC')}",
            styles["Italic"],
        )
    )

    document.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def render_customer_portfolio_xlsx(report: schemas.CustomerPortfolioReport) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Clientes"

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="0f172a")
    data_fill = PatternFill("solid", fgColor="111827")
    accent_fill = PatternFill("solid", fgColor="1d4ed8")

    worksheet["A1"] = "Softmobile 2025 — Reporte de clientes"
    worksheet["A1"].font = Font(color="38bdf8", bold=True, size=16)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    worksheet["A2"] = (
        "Portafolio: clientes morosos"
        if report.category == "delinquent"
        else "Portafolio: clientes frecuentes"
    )
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    worksheet["A3"] = f"Generado: {datetime.utcnow().strftime('%d/%m/%Y %H:%M UTC')}"
    worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)

    summary_headers = ["Indicador", "Valor"]
    summary_rows = [
        ("Clientes incluidos", report.totals.customers),
        ("Marcados morosos", report.totals.moroso_flagged),
        ("Deuda total", _format_currency(report.totals.outstanding_debt)),
        ("Ventas acumuladas", _format_currency(report.totals.sales_total)),
    ]

    worksheet.append(summary_headers)
    header_row = worksheet.max_row
    for column_index in range(1, len(summary_headers) + 1):
        cell = worksheet.cell(row=header_row, column=column_index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for label, value in summary_rows:
        worksheet.append([label, value])
        for column_index in range(1, 3):
            cell = worksheet.cell(row=worksheet.max_row, column=column_index)
            cell.fill = data_fill
            cell.font = Font(color="e2e8f0")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    worksheet.append([])

    headers = [
        "Cliente",
        "Estado",
        "Tipo",
        "Crédito",
        "Saldo",
        "Disponible",
        "Ventas",
        "Operaciones",
        "Última venta",
    ]
    worksheet.append(headers)
    for column_index in range(1, len(headers) + 1):
        cell = worksheet.cell(row=worksheet.max_row, column=column_index)
        cell.font = header_font
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in report.items:
        last_sale_label = ""
        if item.last_sale_at:
            last_sale_label = item.last_sale_at.strftime("%d/%m/%Y %H:%M")
        worksheet.append(
            [
                item.name,
                item.status,
                item.customer_type,
                _format_currency(item.credit_limit),
                _format_currency(item.outstanding_debt),
                _format_currency(item.available_credit),
                _format_currency(item.sales_total),
                item.sales_count,
                last_sale_label,
            ]
        )
        row_index = worksheet.max_row
        for column_index in range(1, len(headers) + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.fill = data_fill
            cell.font = Font(color="e2e8f0")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for column_index in range(1, len(headers) + 1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = 18

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
