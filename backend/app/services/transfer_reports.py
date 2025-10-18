from __future__ import annotations

from collections import Counter
from datetime import datetime
from io import BytesIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .. import models, schemas

DARK_BACKGROUND = colors.HexColor("#0f172a")
DARK_SURFACE = colors.HexColor("#111827")
ACCENT = colors.HexColor("#38bdf8")
TEXT_PRIMARY = colors.HexColor("#e2e8f0")


def _format_timestamp(value: datetime | None) -> str:
    if value is None:
        return "—"
    return value.astimezone().strftime("%d/%m/%Y %H:%M")


def _format_store(store: models.Store | None, store_id: int) -> str:
    if store and getattr(store, "name", None):
        return store.name
    return f"Sucursal #{store_id}"


def build_transfer_report(
    transfers: Sequence[models.TransferOrder],
    filters: schemas.TransferReportFilters,
) -> schemas.TransferReport:
    counters = Counter()
    total_quantity = 0
    items: list[schemas.TransferReportItem] = []

    for transfer in transfers:
        counters[transfer.status.value] += 1
        quantity = sum(item.quantity for item in transfer.items)
        total_quantity += quantity

        devices = [
            schemas.TransferReportDevice(
                sku=item.device.sku if item.device else None,
                name=item.device.name if item.device else None,
                quantity=item.quantity,
            )
            for item in transfer.items
        ]

        folio = f"TR-{transfer.id:06d}"
        if transfer.origin_store and transfer.origin_store.code:
            prefix = transfer.origin_store.code.strip().upper()
            if prefix:
                folio = f"{prefix}-{transfer.id:06d}"

        def _user_label(user: models.User | None) -> str | None:
            if not user:
                return None
            return user.full_name or user.username

        items.append(
            schemas.TransferReportItem(
                id=transfer.id,
                folio=folio,
                origin_store=_format_store(transfer.origin_store, transfer.origin_store_id),
                destination_store=_format_store(
                    transfer.destination_store, transfer.destination_store_id
                ),
                status=transfer.status,
                reason=transfer.reason,
                requested_at=transfer.created_at,
                dispatched_at=transfer.dispatched_at,
                received_at=transfer.received_at,
                cancelled_at=transfer.cancelled_at,
                requested_by=_user_label(transfer.requested_by),
                dispatched_by=_user_label(transfer.dispatched_by),
                received_by=_user_label(transfer.received_by),
                cancelled_by=_user_label(transfer.cancelled_by),
                total_quantity=quantity,
                devices=devices,
            )
        )

    totals = schemas.TransferReportTotals(
        total_transfers=len(transfers),
        pending=counters[models.TransferStatus.SOLICITADA.value],
        in_transit=counters[models.TransferStatus.EN_TRANSITO.value],
        completed=counters[models.TransferStatus.RECIBIDA.value],
        cancelled=counters[models.TransferStatus.CANCELADA.value],
        total_quantity=total_quantity,
    )

    return schemas.TransferReport(
        generated_at=datetime.utcnow(),
        filters=filters,
        totals=totals,
        items=items,
    )


def render_transfer_report_pdf(report: schemas.TransferReport) -> bytes:
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, title="Reporte de transferencias")
    styles = getSampleStyleSheet()

    heading_style = ParagraphStyle(
        "HeadingSoftmobileTransfers",
        parent=styles["Heading1"],
        textColor=ACCENT,
    )
    now_label = datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")

    elements: list = [  # type: ignore[var-annotated]
        Paragraph("Softmobile 2025 — Transferencias entre sucursales", heading_style),
        Spacer(1, 12),
        Paragraph(f"Generado automáticamente el {now_label}", styles["Normal"]),
        Spacer(1, 18),
    ]

    filters_parts: list[str] = []
    if report.filters.store_id:
        filters_parts.append(f"Sucursal monitoreada #{report.filters.store_id}")
    if report.filters.origin_store_id:
        filters_parts.append(f"Origen #{report.filters.origin_store_id}")
    if report.filters.destination_store_id:
        filters_parts.append(f"Destino #{report.filters.destination_store_id}")
    if report.filters.status:
        filters_parts.append(f"Estado: {report.filters.status.value}")
    if report.filters.date_from or report.filters.date_to:
        start = (
            report.filters.date_from.strftime("%d/%m/%Y")
            if report.filters.date_from
            else "∞"
        )
        end = (
            report.filters.date_to.strftime("%d/%m/%Y")
            if report.filters.date_to
            else "∞"
        )
        filters_parts.append(f"Periodo: {start} → {end}")

    filters_label = ", ".join(filters_parts) if filters_parts else "Sin filtros aplicados"
    elements.append(Paragraph(filters_label, styles["Italic"]))
    elements.append(Spacer(1, 12))

    summary_table = Table(
        [
            ["Indicador", "Valor"],
            ["Transferencias", str(report.totals.total_transfers)],
            ["Pendientes", str(report.totals.pending)],
            ["En tránsito", str(report.totals.in_transit)],
            ["Completadas", str(report.totals.completed)],
            ["Canceladas", str(report.totals.cancelled)],
            ["Dispositivos movilizados", str(report.totals.total_quantity)],
        ]
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), DARK_BACKGROUND),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), DARK_SURFACE),
                ("TEXTCOLOR", (0, 1), (-1, -1), TEXT_PRIMARY),
                ("LINEABOVE", (0, 0), (-1, 0), 1, ACCENT),
                ("LINEBELOW", (0, -1), (-1, -1), 1, ACCENT),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ]
        )
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    detail_header = [
        "Folio",
        "Origen",
        "Destino",
        "Estado",
        "Total unidades",
        "Creada",
        "Despachada",
        "Recibida",
    ]
    detail_rows = [detail_header]
    for item in report.items:
        detail_rows.append(
            [
                item.folio,
                item.origin_store,
                item.destination_store,
                item.status.value,
                str(item.total_quantity),
                _format_timestamp(item.requested_at),
                _format_timestamp(item.dispatched_at),
                _format_timestamp(item.received_at),
            ]
        )

    detail_table = Table(detail_rows, hAlign="LEFT")
    detail_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), DARK_BACKGROUND),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), DARK_SURFACE),
                ("TEXTCOLOR", (0, 1), (-1, -1), TEXT_PRIMARY),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#1f2937")),
            ]
        )
    )
    elements.append(detail_table)

    for item in report.items:
        if not item.devices:
            continue
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Detalle de {item.folio}", styles["Heading3"]))
        device_rows = [["SKU", "Descripción", "Cantidad"]]
        for device in item.devices:
            device_rows.append(
                [device.sku or "—", device.name or "Sin descripción", str(device.quantity)]
            )
        devices_table = Table(device_rows, hAlign="LEFT")
        devices_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), DARK_SURFACE),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#0b1120")),
                    ("TEXTCOLOR", (0, 1), (-1, -1), TEXT_PRIMARY),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ]
            )
        )
        elements.append(devices_table)

    document.build(elements)
    return buffer.getvalue()


def render_transfer_report_excel(report: schemas.TransferReport) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transferencias"

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="0f172a")

    sheet.append(
        [
            "Folio",
            "Origen",
            "Destino",
            "Estado",
            "Total unidades",
            "Creada",
            "Despachada",
            "Recibida",
            "Cancelada",
            "Solicitó",
            "Despachó",
            "Recibió",
        ]
    )
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    row_idx = 2
    for item in report.items:
        sheet.append(
            [
                item.folio,
                item.origin_store,
                item.destination_store,
                item.status.value,
                item.total_quantity,
                _format_timestamp(item.requested_at),
                _format_timestamp(item.dispatched_at),
                _format_timestamp(item.received_at),
                _format_timestamp(item.cancelled_at),
                item.requested_by or "—",
                item.dispatched_by or "—",
                item.received_by or "—",
            ]
        )
        row_idx += 1
        if not item.devices:
            continue
        sheet.append(["", "SKU", "Descripción", "Cantidad"])
        row_idx += 1
        for device in item.devices:
            sheet.append([
                "",
                device.sku or "—",
                device.name or "Sin descripción",
                device.quantity,
            ])
            row_idx += 1

    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 20

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
