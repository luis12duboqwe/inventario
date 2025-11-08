"""Renderizadores para los reportes globales (PDF, Excel y CSV)."""
from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .. import schemas
from .global_reports_constants import (
    ACCENT_COLOR,
    ACCENT_COLOR_ARGB,
    GRID_COLOR,
    PRIMARY_BACKGROUND,
    SECONDARY_BACKGROUND,
)

__all__ = [
    "render_global_report_pdf",
    "render_global_report_xlsx",
    "render_global_report_csv",
]


def _format_datetime(value: datetime | None) -> str:
    if value is None:
        return "—"
    return value.strftime("%d/%m/%Y %H:%M UTC")


def _build_pdf_table(data: list[list[str]]) -> Table:
    table = Table(data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY_BACKGROUND)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(SECONDARY_BACKGROUND)),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#e2e8f0")),
                ("LINEABOVE", (0, 0), (-1, 0), 1.2, colors.HexColor(ACCENT_COLOR)),
                ("LINEBELOW", (0, -1), (-1, -1), 1.2, colors.HexColor(ACCENT_COLOR)),
                ("LINEBEFORE", (0, 0), (0, -1), 0.8, colors.HexColor(GRID_COLOR)),
                ("LINEAFTER", (-1, 0), (-1, -1), 0.8, colors.HexColor(GRID_COLOR)),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("LEADING", (0, 0), (-1, -1), 12),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return table


def render_global_report_pdf(
    overview: schemas.GlobalReportOverview,
    dashboard: schemas.GlobalReportDashboard,
) -> bytes:
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, title="Softmobile 2025 — Reporte global")
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "SoftmobileHeading",
        parent=styles["Heading1"],
        textColor=colors.HexColor(ACCENT_COLOR),
        fontSize=18,
    )

    generated_label = overview.generated_at.strftime("%d/%m/%Y %H:%M UTC")
    elements: list = [  # type: ignore[var-annotated]
        Paragraph("Softmobile 2025 — Reporte global de operaciones", title_style),
        Spacer(1, 12),
        Paragraph(f"Generado automáticamente el {generated_label}", styles["Normal"]),
        Spacer(1, 18),
    ]

    filters_parts: list[str] = []
    if overview.filters.date_from:
        filters_parts.append(f"Desde: {_format_datetime(overview.filters.date_from)}")
    if overview.filters.date_to:
        filters_parts.append(f"Hasta: {_format_datetime(overview.filters.date_to)}")
    if overview.filters.module:
        filters_parts.append(f"Módulo: {overview.filters.module}")
    if overview.filters.severity:
        filters_parts.append(f"Severidad: {overview.filters.severity.value}")
    if filters_parts:
        filters_text = " | ".join(filters_parts)
        elements.append(Paragraph(filters_text, styles["Italic"]))
        elements.append(Spacer(1, 12))

    totals_table = _build_pdf_table(
        [
            ["Indicador", "Valor"],
            ["Registros de bitácora", str(overview.totals.logs)],
            ["Errores críticos", str(overview.totals.errors)],
            ["Eventos INFO", str(overview.totals.info)],
            ["Eventos WARNING", str(overview.totals.warning)],
            ["Eventos ERROR", str(overview.totals.error)],
            ["Eventos CRITICAL", str(overview.totals.critical)],
            ["Sync pendientes", str(overview.totals.sync_pending)],
            ["Sync fallidas", str(overview.totals.sync_failed)],
            ["Última actividad", _format_datetime(overview.totals.last_activity_at)],
        ]
    )
    elements.append(totals_table)
    elements.append(Spacer(1, 18))

    if overview.module_breakdown:
        module_data: list[list[str]] = [["Módulo", "Eventos"]]
        module_data.extend(
            [[item.name.title(), str(item.total)] for item in overview.module_breakdown]
        )
        elements.append(Paragraph("Distribución por módulo", styles["Heading3"]))
        elements.append(Spacer(1, 6))
        elements.append(_build_pdf_table(module_data))
        elements.append(Spacer(1, 18))

    if overview.severity_breakdown:
        severity_data: list[list[str]] = [["Severidad", "Eventos"]]
        severity_data.extend(
            [[item.name.upper(), str(item.total)] for item in overview.severity_breakdown]
        )
        elements.append(Paragraph("Distribución por severidad", styles["Heading3"]))
        elements.append(Spacer(1, 6))
        elements.append(_build_pdf_table(severity_data))
        elements.append(Spacer(1, 18))

    if dashboard.activity_series:
        activity_rows: list[list[str]] = [[
            "Fecha",
            "Info",
            "Warning",
            "Error",
            "Critical",
            "Errores sistema",
        ]]
        for point in dashboard.activity_series[-14:]:
            activity_rows.append(
                [
                    point.date.strftime("%d/%m/%Y"),
                    str(point.info),
                    str(point.warning),
                    str(point.error),
                    str(point.critical),
                    str(point.system_errors),
                ]
            )
        elements.append(Paragraph("Actividad reciente", styles["Heading3"]))
        elements.append(Spacer(1, 6))
        elements.append(_build_pdf_table(activity_rows))
        elements.append(Spacer(1, 18))

    if overview.alerts:
        alert_rows: list[list[str]] = [
            ["Tipo", "Severidad", "Módulo", "Mensaje", "Eventos", "Último evento"],
        ]
        for alert in overview.alerts[:15]:
            alert_rows.append(
                [
                    alert.type.replace("_", " ").title(),
                    alert.level.value.upper(),
                    alert.module or "general",
                    alert.message,
                    str(alert.count),
                    _format_datetime(alert.occurred_at),
                ]
            )
        elements.append(Paragraph("Alertas automáticas", styles["Heading3"]))
        elements.append(Spacer(1, 6))
        elements.append(_build_pdf_table(alert_rows))
        elements.append(Spacer(1, 18))

    if overview.recent_logs:
        log_rows: list[list[str]] = [["Fecha", "Usuario", "Módulo", "Acción", "Nivel"]]
        for entry in overview.recent_logs[:15]:
            log_rows.append(
                [
                    _format_datetime(entry.fecha),
                    entry.usuario or "Sistema",
                    entry.modulo,
                    entry.accion,
                    entry.nivel.value,
                ]
            )
        elements.append(Paragraph("Actividad reciente", styles["Heading3"]))
        elements.append(Spacer(1, 6))
        elements.append(_build_pdf_table(log_rows))
        elements.append(Spacer(1, 18))

    if overview.recent_errors:
        error_rows: list[list[str]] = [["Fecha", "Módulo", "Mensaje", "Usuario"]]
        for entry in overview.recent_errors[:15]:
            error_rows.append(
                [
                    _format_datetime(entry.fecha),
                    entry.modulo,
                    entry.mensaje,
                    entry.usuario or "Sistema",
                ]
            )
        elements.append(Paragraph("Errores del sistema", styles["Heading3"]))
        elements.append(Spacer(1, 6))
        elements.append(_build_pdf_table(error_rows))

    document.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def render_global_report_xlsx(
    overview: schemas.GlobalReportOverview,
    dashboard: schemas.GlobalReportDashboard,
) -> BytesIO:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"

    header_font = Font(color="FFFFFF", bold=True)
    accent_font = Font(color=ACCENT_COLOR_ARGB, bold=True, size=16)
    header_fill = PatternFill("solid", fgColor="0f172a")
    accent_fill = PatternFill("solid", fgColor="1d4ed8")
    data_fill = PatternFill("solid", fgColor="111827")

    summary_sheet["A1"] = "Softmobile 2025 — Reporte global"
    summary_sheet["A1"].font = accent_font
    summary_sheet.merge_cells("A1:F1")

    summary_sheet["A2"] = f"Generado: {overview.generated_at.strftime('%d/%m/%Y %H:%M UTC')}"
    summary_sheet.merge_cells("A2:F2")

    filters_parts: list[str] = []
    if overview.filters.date_from:
        filters_parts.append(f"Desde {overview.filters.date_from.strftime('%d/%m/%Y %H:%M')} UTC")
    if overview.filters.date_to:
        filters_parts.append(f"Hasta {overview.filters.date_to.strftime('%d/%m/%Y %H:%M')} UTC")
    if overview.filters.module:
        filters_parts.append(f"Módulo: {overview.filters.module}")
    if overview.filters.severity:
        filters_parts.append(f"Severidad: {overview.filters.severity.value}")
    if filters_parts:
        summary_sheet["A3"] = " | ".join(filters_parts)
        summary_sheet.merge_cells("A3:F3")

    summary_sheet.append([])
    summary_sheet.append(["Indicador", "Valor"])
    header_row = summary_sheet.max_row
    for column in range(1, 3):
        cell = summary_sheet.cell(row=header_row, column=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    metric_rows = [
        ("Registros de bitácora", overview.totals.logs),
        ("Errores críticos", overview.totals.errors),
        ("Eventos INFO", overview.totals.info),
        ("Eventos WARNING", overview.totals.warning),
        ("Eventos ERROR", overview.totals.error),
        ("Eventos CRITICAL", overview.totals.critical),
        ("Sync pendientes", overview.totals.sync_pending),
        ("Sync fallidas", overview.totals.sync_failed),
        ("Última actividad", _format_datetime(overview.totals.last_activity_at)),
    ]
    for label, value in metric_rows:
        summary_sheet.append([label, value])
        row_index = summary_sheet.max_row
        for column in range(1, 3):
            cell = summary_sheet.cell(row=row_index, column=column)
            cell.fill = data_fill
            cell.font = Font(color="e2e8f0")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    summary_sheet.append([])

    if overview.module_breakdown:
        summary_sheet.append(["Módulo", "Eventos"])
        header_row = summary_sheet.max_row
        for column in range(1, 3):
            cell = summary_sheet.cell(row=header_row, column=column)
            cell.font = header_font
            cell.fill = accent_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for item in overview.module_breakdown:
            summary_sheet.append([item.name, item.total])
            row_index = summary_sheet.max_row
            for column in range(1, 3):
                cell = summary_sheet.cell(row=row_index, column=column)
                cell.fill = data_fill
                cell.font = Font(color="e2e8f0")
                cell.alignment = Alignment(horizontal="left", vertical="center")
        summary_sheet.append([])

    if overview.severity_breakdown:
        summary_sheet.append(["Severidad", "Eventos"])
        header_row = summary_sheet.max_row
        for column in range(1, 3):
            cell = summary_sheet.cell(row=header_row, column=column)
            cell.font = header_font
            cell.fill = accent_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for item in overview.severity_breakdown:
            summary_sheet.append([item.name, item.total])
            row_index = summary_sheet.max_row
            for column in range(1, 3):
                cell = summary_sheet.cell(row=row_index, column=column)
                cell.fill = data_fill
                cell.font = Font(color="e2e8f0")
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for column in range(1, 7):
        summary_sheet.column_dimensions[get_column_letter(column)].width = 22

    activity_sheet = workbook.create_sheet("Actividad")
    activity_sheet.append([
        "Fecha",
        "Info",
        "Warning",
        "Error",
        "Critical",
        "Errores sistema",
    ])
    header_row = activity_sheet.max_row
    for column in range(1, 7):
        cell = activity_sheet.cell(row=header_row, column=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for point in dashboard.activity_series:
        activity_sheet.append(
            [
                point.date.strftime("%d/%m/%Y"),
                point.info,
                point.warning,
                point.error,
                point.critical,
                point.system_errors,
            ]
        )
    for column in range(1, 7):
        activity_sheet.column_dimensions[get_column_letter(column)].width = 18

    alerts_sheet = workbook.create_sheet("Alertas")
    alerts_sheet.append([
        "Tipo",
        "Severidad",
        "Módulo",
        "Mensaje",
        "Eventos",
        "Último evento",
    ])
    header_row = alerts_sheet.max_row
    for column in range(1, 7):
        cell = alerts_sheet.cell(row=header_row, column=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for alert in overview.alerts:
        alerts_sheet.append(
            [
                alert.type,
                alert.level.value,
                alert.module or "general",
                alert.message,
                alert.count,
                _format_datetime(alert.occurred_at),
            ]
        )
    for column in range(1, 7):
        alerts_sheet.column_dimensions[get_column_letter(column)].width = 28

    logs_sheet = workbook.create_sheet("Logs")
    logs_sheet.append(["Fecha", "Usuario", "Módulo", "Acción", "Nivel", "Descripción"])
    header_row = logs_sheet.max_row
    for column in range(1, 7):
        cell = logs_sheet.cell(row=header_row, column=column)
        cell.font = header_font
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for entry in overview.recent_logs:
        logs_sheet.append(
            [
                _format_datetime(entry.fecha),
                entry.usuario or "Sistema",
                entry.modulo,
                entry.accion,
                entry.nivel.value,
                entry.descripcion,
            ]
        )
    for column in range(1, 7):
        logs_sheet.column_dimensions[get_column_letter(column)].width = 26

    errors_sheet = workbook.create_sheet("Errores")
    errors_sheet.append(["Fecha", "Módulo", "Mensaje", "Usuario"])
    header_row = errors_sheet.max_row
    for column in range(1, 5):
        cell = errors_sheet.cell(row=header_row, column=column)
        cell.font = header_font
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for entry in overview.recent_errors:
        errors_sheet.append(
            [
                _format_datetime(entry.fecha),
                entry.modulo,
                entry.mensaje,
                entry.usuario or "Sistema",
            ]
        )
    for column in range(1, 5):
        errors_sheet.column_dimensions[get_column_letter(column)].width = 32

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def render_global_report_csv(
    overview: schemas.GlobalReportOverview,
    dashboard: schemas.GlobalReportDashboard,
) -> StringIO:
    buffer = StringIO()
    writer = csv.writer(buffer)

    writer.writerow(["Softmobile 2025 — Reporte global de operaciones"])
    writer.writerow([f"Generado", overview.generated_at.strftime("%d/%m/%Y %H:%M UTC")])
    if overview.filters.date_from:
        writer.writerow(["Desde", overview.filters.date_from.isoformat()])
    if overview.filters.date_to:
        writer.writerow(["Hasta", overview.filters.date_to.isoformat()])
    if overview.filters.module:
        writer.writerow(["Módulo", overview.filters.module])
    if overview.filters.severity:
        writer.writerow(["Severidad", overview.filters.severity.value])

    writer.writerow([])
    writer.writerow(["Indicador", "Valor"])
    writer.writerow(["Registros", overview.totals.logs])
    writer.writerow(["Errores críticos", overview.totals.errors])
    writer.writerow(["Eventos INFO", overview.totals.info])
    writer.writerow(["Eventos WARNING", overview.totals.warning])
    writer.writerow(["Eventos ERROR", overview.totals.error])
    writer.writerow(["Eventos CRITICAL", overview.totals.critical])
    writer.writerow(["Sync pendientes", overview.totals.sync_pending])
    writer.writerow(["Sync fallidas", overview.totals.sync_failed])
    writer.writerow(["Última actividad", _format_datetime(overview.totals.last_activity_at)])

    if overview.module_breakdown:
        writer.writerow([])
        writer.writerow(["Distribución por módulo"])
        writer.writerow(["Módulo", "Eventos"])
        for item in overview.module_breakdown:
            writer.writerow([item.name, item.total])

    if overview.severity_breakdown:
        writer.writerow([])
        writer.writerow(["Distribución por severidad"])
        writer.writerow(["Severidad", "Eventos"])
        for item in overview.severity_breakdown:
            writer.writerow([item.name, item.total])

    if dashboard.activity_series:
        writer.writerow([])
        writer.writerow(["Actividad reciente"])
        writer.writerow(["Fecha", "Info", "Warning", "Error", "Critical", "Errores sistema"])
        for point in dashboard.activity_series:
            writer.writerow(
                [
                    point.date.isoformat(),
                    point.info,
                    point.warning,
                    point.error,
                    point.critical,
                    point.system_errors,
                ]
            )

    if overview.alerts:
        writer.writerow([])
        writer.writerow(["Alertas"])
        writer.writerow(["Tipo", "Severidad", "Módulo", "Mensaje", "Eventos", "Último evento"])
        for alert in overview.alerts:
            writer.writerow(
                [
                    alert.type,
                    alert.level.value,
                    alert.module or "general",
                    alert.message,
                    alert.count,
                    _format_datetime(alert.occurred_at),
                ]
            )

    if overview.recent_logs:
        writer.writerow([])
        writer.writerow(["Logs recientes"])
        writer.writerow(["Fecha", "Usuario", "Módulo", "Acción", "Nivel", "Descripción"])
        for entry in overview.recent_logs:
            writer.writerow(
                [
                    entry.fecha.isoformat(),
                    entry.usuario or "Sistema",
                    entry.modulo,
                    entry.accion,
                    entry.nivel.value,
                    entry.descripcion,
                ]
            )

    if overview.recent_errors:
        writer.writerow([])
        writer.writerow(["Errores del sistema"])
        writer.writerow(["Fecha", "Módulo", "Mensaje", "Usuario"])
        for entry in overview.recent_errors:
            writer.writerow(
                [
                    entry.fecha.isoformat(),
                    entry.modulo,
                    entry.mensaje,
                    entry.usuario or "Sistema",
                ]
            )

    buffer.seek(0)
    return buffer
