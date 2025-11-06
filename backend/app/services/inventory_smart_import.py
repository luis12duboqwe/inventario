from __future__ import annotations

"""Herramientas para la importación inteligente de inventario.

Este módulo concentra la lectura flexible de archivos tabulares, la detección de
encabezados y la persistencia segura de dispositivos con tolerancia a datos
incompletos. La intención de los docstrings añadidos es documentar los pasos del
análisis y los puntos de extensión usados por los equipos de catálogo y
sincronización híbrida.
"""

import csv
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from time import perf_counter
from typing import Any, Iterable
from zipfile import BadZipFile

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..core.transactions import transactional_session
from .. import crud, models, schemas
from . import import_validation

IMEI_PATTERN = re.compile(r"\b\d{15}\b")

ESTADOS_COMERCIALES_VALIDOS: set[str] = {
    estado.name for estado in models.CommercialState
} | {estado.value.upper() for estado in models.CommercialState}

ESTADO_COMERCIAL_FIX_SUGERIDO = models.CommercialState.NUEVO.value.upper()

BOOLEAN_TRUE_VALUES = frozenset(
    {
        "si",
        "sí",
        "yes",
        "true",
        "verdadero",
        "activo",
        "habilitado",
        "1",
    }
)
BOOLEAN_FALSE_VALUES = frozenset(
    {
        "no",
        "false",
        "falso",
        "inactivo",
        "deshabilitado",
        "0",
    }
)

CANONICAL_FIELDS: dict[str, set[str]] = {
    "sku": {"sku", "codigo", "product_code", "code"},
    "name": {"nombre", "name", "descripcion", "description", "producto"},
    "marca": {"marca", "brand", "fabricante", "manufacturer"},
    "modelo": {"modelo", "model", "device", "modelo_equipo"},
    "imei": {"imei", "imei1", "imei_1", "imei principal"},
    "serial": {"serial", "serie", "numero_serie", "serial_number"},
    "color": {"color", "colour"},
    "capacidad": {"capacidad", "storage", "memoria", "ram"},
    "capacidad_gb": {"capacidad_gb", "gb", "storage_gb"},
    "tienda": {"tienda", "sucursal", "store", "shop", "ubicacion", "branch"},
    "estado": {"estado", "status", "estado_actual", "situacion"},
    "categoria": {"categoria", "category", "segmento"},
    "condicion": {"condicion", "condition", "grado"},
    "cantidad": {"cantidad", "qty", "existencia", "stock", "units"},
    "precio": {"precio", "precio_venta", "precio_publico", "price", "unit_price"},
    "costo": {"costo", "costo_compra", "cost", "unit_cost"},
    "ubicacion": {"ubicacion_interna", "pasillo", "rack", "ubicacion"},
    "proveedor": {"proveedor", "vendor", "supplier"},
    "lote": {"lote", "batch"},
    "fecha_compra": {"fecha_compra", "purchase_date", "fecha de compra"},
    "fecha_ingreso": {"fecha_ingreso", "arrival_date", "fecha alta"},
    "estado_comercial": {"estado_comercial", "grade", "tier"},
}

CRITICAL_FIELDS = {"marca", "modelo", "tienda"}
IMEI_IMPORTANT_FIELDS = {"imei"}


@dataclass
class ParsedFile:
    headers: list[str]
    rows: list[dict[str, Any]]


def process_smart_import(
    db: Session,
    *,
    file_bytes: bytes,
    filename: str,
    commit: bool,
    overrides: dict[str, str] | None,
    performed_by_id: int | None,
    username: str | None,
    reason: str,
) -> schemas.InventorySmartImportResponse:
    """Analiza o persiste un archivo tabular de inventario.

    El flujo se compone de tres pasos principales:

    1. ``_read_tabular_file`` interpreta CSV o Excel (con *fallback* a CSV si el
       archivo está dañado) y genera una estructura homogénea ``ParsedFile``.
    2. ``_analyze_dataset`` cruza encabezados con sinónimos conocidos y
       aprendizaje previo para construir la vista previa consumida por la UI.
    3. ``_commit_import`` crea/actualiza dispositivos y registra historial sólo
       cuando ``commit`` es ``True``. Esta fase se ejecuta dentro de una
       transacción para mantener la consistencia entre dispositivos y bitácoras.

    Los docstrings detallan dependencias y escenarios cubiertos por
    ``tests/test_inventory_smart_import.py`` para facilitar extensiones futuras.
    """
    overrides = overrides or {}
    parsed = _read_tabular_file(file_bytes, filename)
    learned_patterns = crud.get_known_import_column_patterns(db)
    preview = _analyze_dataset(
        db, parsed, overrides=overrides, learned_patterns=learned_patterns)
    # Normalización proactiva: garantizar que sucursales creadas durante la vista previa
    # (cuando se confirme posteriormente) cuenten con timezone corporativo por defecto y
    # inventory_value inicializado en 0. Esta lógica se ejecuta sólo en commit dentro de
    # _commit_import, pero dejamos aquí un comentario explícito para la traza operativa.
    result: schemas.InventorySmartImportResult | None = None
    if commit:
        result = _commit_import(
            db,
            parsed,
            preview,
            filename=filename,
            performed_by_id=performed_by_id,
            username=username,
            overrides=overrides,
            reason=reason,
        )
    return schemas.InventorySmartImportResponse(preview=preview, resultado=result)


def _parse_csv_bytes(file_bytes: bytes) -> ParsedFile:
    """Convierte ``bytes`` CSV en ``ParsedFile`` normalizado.

    Asegura la limpieza de encabezados y descarta filas completamente vacías para
    evitar falsos positivos durante el análisis y la escritura en base.
    """
    decoded = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(decoded))
    if reader.fieldnames is None:
        raise ValueError("archivo_sin_encabezados")
    headers = [header.strip() for header in reader.fieldnames]
    rows: list[dict[str, Any]] = []
    for raw in reader:
        normalized_row = {(key or "").strip(): value for key,
                          value in raw.items()}
        if not _row_has_data(normalized_row.values()):
            continue
        rows.append(normalized_row)
    return ParsedFile(headers=headers, rows=rows)


def _read_tabular_file(file_bytes: bytes, filename: str) -> ParsedFile:
    """Lee archivos CSV/XLSX devolviendo filas homogéneas para el análisis.

    La función intenta abrir Excel en modo de sólo lectura; si el contenedor ZIP
    es inválido se vuelve a intentar como CSV. Los encabezados vacíos provocan un
    ``ValueError`` que es propagado hasta la capa API para mostrar mensajes en la
    UI de importaciones inteligentes.
    """
    if filename.lower().endswith(".csv"):
        return _parse_csv_bytes(file_bytes)
    # default excel
    try:
        workbook = load_workbook(
            BytesIO(file_bytes), read_only=True, data_only=True)
    except BadZipFile:
        return _parse_csv_bytes(file_bytes)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:  # pragma: no cover - empty file defensive
        raise ValueError("archivo_vacio") from exc
    headers = [str(cell).strip()
               if cell is not None else "" for cell in header_row]
    cleaned_headers = [header for header in headers if header]
    if not cleaned_headers:
        raise ValueError("archivo_sin_encabezados")
    parsed_rows: list[dict[str, Any]] = []
    for row in rows_iter:
        record = {headers[index]: row[index]
                  for index in range(min(len(headers), len(row)))}
        if not _row_has_data(record.values()):
            continue
        parsed_rows.append(record)
    return ParsedFile(headers=headers, rows=parsed_rows)


def _analyze_dataset(
    db: Session,
    parsed: ParsedFile,
    *,
    overrides: dict[str, str],
    learned_patterns: dict[str, str],
) -> schemas.InventorySmartImportPreview:
    """Construye la vista previa de importación con base en sinónimos y muestras.

    - Asocia encabezados a campos canónicos combinando sinónimos, *overrides* y
      patrones aprendidos mediante ``crud.get_known_import_column_patterns``.
    - Calcula advertencias de columnas faltantes o no mapeadas y estima registros
      incompletos según ``CRITICAL_FIELDS`` y ``IMEI_IMPORTANT_FIELDS``.
    - Prepara ``SmartImportColumnMatch`` consumidos por el frontend para mostrar
      estados («ok»/«pendiente»/«falta»).

    Las decisiones documentadas están cubiertas por ``test_inventory_smart_import_preview_and_commit``
    y ``test_inventory_smart_import_handles_overrides_and_incomplete_records``.
    """
    normalized_headers = {
        _normalize_header(header): header for header in parsed.headers if header
    }
    column_samples: dict[str, list[str]] = defaultdict(list)
    for row in parsed.rows:
        for original_header, value in row.items():
            if not original_header:
                continue
            normalized_value = _normalize_cell(value)
            if normalized_value is not None:
                column_samples[original_header].append(normalized_value)
    column_map: dict[str, str | None] = {}
    patrones_sugeridos: dict[str, str] = {}
    warnings: list[str] = []

    for canonical, synonyms in CANONICAL_FIELDS.items():
        mapped_header = _resolve_header(
            canonical,
            synonyms,
            normalized_headers,
            overrides,
            learned_patterns,
            column_samples,
        )
        if mapped_header:
            column_map[canonical] = mapped_header
            patrones_sugeridos[_normalize_header(mapped_header)] = canonical
        else:
            column_map[canonical] = None

    missing_columns = [field for field,
                       header in column_map.items() if header is None]
    if missing_columns:
        warnings.append(
            "Columnas faltantes: " + ", ".join(sorted(missing_columns))
        )

    unknown_headers = [
        header
        for normalized, header in normalized_headers.items()
        if normalized not in patrones_sugeridos
    ]
    if unknown_headers:
        warnings.append(
            "Columnas sin asignar detectadas: " +
            ", ".join(sorted(set(unknown_headers)))
        )

    registros_incompletos = 0
    total_rows = len(parsed.rows)
    column_matches: list[schemas.SmartImportColumnMatch] = []
    canonical_values_per_row = _extract_canonical_rows(parsed.rows, column_map)
    for canonical, header in column_map.items():
        if header is None:
            column_matches.append(
                schemas.SmartImportColumnMatch(
                    campo=canonical,
                    encabezado_origen=None,
                    estado="falta",
                    tipo_dato=None,
                    ejemplos=[],
                )
            )
            continue
        samples = column_samples.get(header, [])
        filled = len(samples)
        estado = "ok" if filled == total_rows and total_rows > 0 else "pendiente"
        detected_type = _detect_column_type(samples)
        column_matches.append(
            schemas.SmartImportColumnMatch(
                campo=canonical,
                encabezado_origen=header,
                estado=estado,
                tipo_dato=detected_type,
                ejemplos=samples[:3],
            )
        )

    for canonical_row in canonical_values_per_row:
        if _is_row_incomplete(canonical_row):
            registros_incompletos += 1

    return schemas.InventorySmartImportPreview(
        columnas=column_matches,
        columnas_detectadas=column_map,
        columnas_faltantes=missing_columns,
        total_filas=total_rows,
        registros_incompletos_estimados=registros_incompletos,
        advertencias=warnings,
        patrones_sugeridos=patrones_sugeridos,
    )


def _commit_import(
    db: Session,
    parsed: ParsedFile,
    preview: schemas.InventorySmartImportPreview,
    *,
    filename: str,
    performed_by_id: int | None,
    username: str | None,
    overrides: dict[str, str],
    reason: str,
) -> schemas.InventorySmartImportResult:
    """Persiste dispositivos y registra historial cuando la importación es confirmada.

    La rutina recorre ``canonical_rows`` y aplica la siguiente estrategia:

    * Garantiza la existencia de la sucursal con ``crud.ensure_store_by_name``
      antes de crear dispositivos.
    * Calcula *payloads* a partir de campos canónicos, normalizando cantidades,
      fechas y precios mediante los helpers ``_parse_int`` y ``_parse_decimal``.
    * Identifica registros incompletos y agrega notas a ``import_validation`` para
      que el endpoint ``/inventory/devices/incomplete`` pueda exhibirlos.
    * Registra la operación en el historial con duración y advertencias para la
      auditoría de inventario.

    ``test_inventory_smart_import_preview_and_commit`` y
    ``test_inventory_smart_import_handles_overrides_and_incomplete_records``
    validan los escenarios críticos descritos en esta documentación.
    """
    start_time = perf_counter()
    canonical_rows = _extract_canonical_rows(
        parsed.rows, preview.columnas_detectadas)
    total_processed = 0
    created = 0
    updated = 0
    registros_incompletos = 0
    warnings = list(preview.advertencias)
    new_stores: list[str] = []
    processed_records: list[dict[str, Any]] = []
    estado_comercial_incidencias: list[
        import_validation.CommercialStateIncident
    ] = []
    duration = 0.0
    resumen = ""
    resumen_validacion: schemas.ImportValidationSummary | None = None
    with transactional_session(db):
        for row_index, row in enumerate(canonical_rows, start=1):
            total_processed += 1
            raw_quantity = row.get("cantidad")
            parsed_quantity = _parse_int(raw_quantity)
            quantity = (
                parsed_quantity if parsed_quantity is not None and parsed_quantity >= 0 else 0
            )
            raw_costo = row.get("costo")
            costo = _parse_decimal(raw_costo)
            raw_precio = row.get("precio")
            precio = _parse_decimal(raw_precio)
            fecha_compra = _parse_date(row.get("fecha_compra"))
            fecha_ingreso = _parse_date(row.get("fecha_ingreso"))
            imei = _validate_imei(row.get("imei"))
            serial = _normalize_optional(row.get("serial"))
            store_name = row.get("tienda")
            record_kwargs: dict[str, Any] = {
                "row_index": row_index,
                "store_id": None,
                "store_name": store_name,
                "imei": imei,
                "serial": serial,
                "raw_cantidad": raw_quantity,
                "parsed_cantidad": parsed_quantity,
                "raw_precio": raw_precio,
                "parsed_precio": precio,
                "raw_costo": raw_costo,
                "parsed_costo": costo,
                "fecha_compra": fecha_compra,
                "fecha_ingreso": fecha_ingreso,
                "device_id": None,
            }
            if not store_name:
                warnings.append(
                    f"Fila {row_index}: no se especificó la tienda.")
                registros_incompletos += 1
                processed_records.append(
                    import_validation.build_record(**record_kwargs))
                continue
            store, was_created = crud.ensure_store_by_name(
                db,
                store_name,
                performed_by_id=performed_by_id,
            )
            # Hook de normalización de sucursal recién creada: timezone y valor inventario.
            if was_created:
                # Alinear timezone estándar corporativo si quedó en blanco o genérico.
                if not store.timezone or store.timezone == "UTC":
                    store.timezone = "America/Mexico_City"
                # inventory_value debe mantenerse en 0 Decimal de forma explícita
                # para instalaciones con motores que no respetan defaults al crear via ensure_store_by_name.
                if store.inventory_value is None:
                    from decimal import Decimal as _D  # import local para aislar dependencia
                    store.inventory_value = _D("0")
                db.add(store)
            record_kwargs["store_id"] = store.id
            record_kwargs["store_name"] = store.name
            if was_created:
                new_stores.append(store.name)
            sku = row.get("sku") or _generate_sku(
                store.code, row.get("modelo"), row_index)
            name = row.get("name") or _generate_name(row)
            capacidad_gb = _parse_int(row.get("capacidad_gb"))
            estado = row.get("estado") or "pendiente"
            estado_comercial, estado_comercial_original = _resolve_estado_comercial(
                row.get("estado_comercial")
            )
            completo = not _is_row_incomplete(row)
            if not completo:
                registros_incompletos += 1
            base_payload = {
                "sku": sku,
                "name": name,
                "quantity": quantity,
                "unit_price": precio or Decimal("0"),
                "precio_venta": precio or Decimal("0"),
                "costo_unitario": costo or Decimal("0"),
                "costo_compra": costo or Decimal("0"),
                "marca": row.get("marca"),
                "modelo": row.get("modelo"),
                "color": row.get("color"),
                "capacidad": row.get("capacidad"),
                "capacidad_gb": capacidad_gb,
                "estado": estado,
                "estado_comercial": estado_comercial,
                "categoria": row.get("categoria"),
                "condicion": row.get("condicion"),
                "ubicacion": row.get("ubicacion"),
                "proveedor": row.get("proveedor"),
                "lote": row.get("lote"),
                "descripcion": row.get("descripcion") or name,
                "imagen_url": None,
                "imei": imei,
                "serial": serial,
                "fecha_compra": fecha_compra,
                "fecha_ingreso": fecha_ingreso,
                "completo": completo,
            }
            existing = crud.find_device_for_import(
                db,
                store_id=store.id,
                imei=imei,
                serial=serial,
                modelo=row.get("modelo"),
                color=row.get("color"),
            )
            if existing and existing.store_id != store.id:
                warnings.append(
                    f"Fila {row_index}: el dispositivo con identificadores coincide con otra sucursal. Se omite la actualización."
                )
                registros_incompletos += 1
                record_kwargs["device_id"] = existing.id
                processed_records.append(
                    import_validation.build_record(**record_kwargs))
                continue
            device: models.Device | None = existing
            if existing is None:
                payload = schemas.DeviceCreate(**base_payload)
                try:
                    device = crud.create_device(
                        db,
                        store.id,
                        payload,
                        performed_by_id=performed_by_id,
                    )
                except ValueError as exc:  # pragma: no cover - defensive against race conditions
                    warnings.append(
                        f"Fila {row_index}: no se pudo crear el dispositivo ({exc})."
                    )
                    registros_incompletos += 1
                    processed_records.append(
                        import_validation.build_record(**record_kwargs))
                    continue
                created += 1
            else:
                update_payload = {
                    key: value
                    for key, value in base_payload.items()
                    if key
                    not in {
                        "sku",
                        "quantity",
                        "unit_price",
                        "precio_venta",
                        "costo_unitario",
                        "costo_compra",
                    }
                    and value is not None
                }
                update_numeric: dict[str, Any] = {}
                update_numeric["quantity"] = quantity
                if precio is not None:
                    update_numeric["unit_price"] = precio
                    update_numeric["precio_venta"] = precio
                if costo is not None:
                    update_numeric["costo_unitario"] = costo
                    update_numeric["costo_compra"] = costo
                update_payload.update(update_numeric)
                update_payload["completo"] = completo
                try:
                    device = crud.update_device(
                        db,
                        store.id,
                        existing.id,
                        schemas.DeviceUpdate(**update_payload),
                        performed_by_id=performed_by_id,
                    )
                except ValueError as exc:  # pragma: no cover - should be rare
                    warnings.append(
                        f"Fila {row_index}: no se pudo actualizar el dispositivo ({exc})."
                    )
                    registros_incompletos += 1
                    processed_records.append(
                        import_validation.build_record(**record_kwargs))
                    continue
                updated += 1
            record_kwargs["device_id"] = device.id if device else None
            if estado_comercial_original:
                estado_comercial_incidencias.append(
                    {
                        "row_index": row_index,
                        "device_id": record_kwargs["device_id"],
                        "valor_original": estado_comercial_original,
                        "fix_sugerido": ESTADO_COMERCIAL_FIX_SUGERIDO,
                    }
                )
            if quantity is not None and quantity >= 0 and device is not None:
                movement_payload = schemas.MovementCreate(
                    producto_id=device.id,
                    tipo_movimiento=models.MovementType.ADJUST,
                    cantidad=quantity,
                    comentario="Importación inteligente v2.2.0",
                    sucursal_origen_id=None,
                    sucursal_destino_id=None,
                    unit_cost=costo,
                )
                try:
                    crud.create_inventory_movement(
                        db,
                        store.id,
                        movement_payload,
                        performed_by_id=performed_by_id,
                    )
                except ValueError as exc:
                    warnings.append(
                        f"Fila {row_index}: no se pudo registrar el movimiento ({exc})."
                    )
            processed_records.append(
                import_validation.build_record(**record_kwargs))
        duration = perf_counter() - start_time
        warnings = list(dict.fromkeys(warnings))
        resumen = (
            "📦 Resultado de importación:\n"
            f"- Total procesados: {total_processed}\n"
            f"- Nuevos productos: {created}\n"
            f"- Actualizados: {updated}\n"
            f"- Columnas faltantes: {len(preview.columnas_faltantes)} ({', '.join(preview.columnas_faltantes) if preview.columnas_faltantes else '0'})\n"
            f"- Registros incompletos: {registros_incompletos}\n"
            f"- Tiendas nuevas: {len(new_stores)}\n"
            f"- Motivo: {reason}"
        )
        crud.create_inventory_import_record(
            db,
            filename=filename,
            columnas_detectadas=preview.columnas_detectadas,
            registros_incompletos=registros_incompletos,
            total_registros=total_processed,
            nuevos=created,
            actualizados=updated,
            advertencias=warnings,
            patrones_columnas=preview.patrones_sugeridos,
            duration_seconds=duration,
        )
        resumen_validacion = import_validation.validar_importacion(
            db,
            registros=processed_records,
            columnas_faltantes=preview.columnas_faltantes,
            import_duration=duration,
            incidencias_estado_comercial=estado_comercial_incidencias,
        )
        crud._create_system_log(  # type: ignore[attr-defined]
            db,
            audit_log=None,
            usuario=username,
            module="inventario",
            action="inventory_smart_import",
            description=(
                f"Importación inteligente ejecutada sobre {filename}: "
                f"{total_processed} filas, {created} nuevas, {updated} actualizadas, {registros_incompletos} incompletas"
                f". Motivo: {reason}"
            ),
            level=models.SystemLogLevel.INFO,
        )
        for warning in warnings:
            crud._create_system_log(  # type: ignore[attr-defined]
                db,
                audit_log=None,
                usuario=username,
                module="inventario",
                action="inventory_smart_import_warning",
                description=warning,
                level=models.SystemLogLevel.WARNING,
            )
    return schemas.InventorySmartImportResult(
        total_procesados=total_processed,
        nuevos=created,
        actualizados=updated,
        registros_incompletos=registros_incompletos,
        columnas_faltantes=preview.columnas_faltantes,
        advertencias=warnings,
        tiendas_nuevas=sorted(set(new_stores)),
        duracion_segundos=round(duration, 2),
        resumen=resumen,
        validacion_resumen=resumen_validacion,
    )


def _resolve_header(
    canonical: str,
    synonyms: Iterable[str],
    normalized_headers: dict[str, str],
    overrides: dict[str, str],
    learned_patterns: dict[str, str],
    samples: dict[str, list[str]],
) -> str | None:
    override = overrides.get(canonical)
    if override:
        normalized_override = _normalize_header(override)
        if normalized_override in normalized_headers:
            return normalized_headers[normalized_override]
    for synonym in synonyms:
        normalized_synonym = _normalize_header(synonym)
        if normalized_synonym in normalized_headers:
            return normalized_headers[normalized_synonym]
    for normalized_header, canonical_target in learned_patterns.items():
        if canonical_target == canonical and normalized_header in normalized_headers:
            return normalized_headers[normalized_header]
    if canonical == "imei":
        for normalized, header in normalized_headers.items():
            values = samples.get(header, [])
            if values and all(IMEI_PATTERN.fullmatch(value) for value in values[:5]):
                return header
    if canonical == "cantidad":
        for normalized, header in normalized_headers.items():
            values = samples.get(header, [])
            if values and all(value.replace(".", "", 1).isdigit() for value in values[:5]):
                if any(keyword in normalized for keyword in {"cantidad", "qty", "stock", "existencia"}):
                    return header
    return None


def _extract_canonical_rows(
    rows: list[dict[str, Any]],
    column_map: dict[str, str | None],
) -> list[dict[str, Any]]:
    canonical_rows: list[dict[str, Any]] = []
    for raw in rows:
        canonical_row: dict[str, Any] = {}
        for canonical, header in column_map.items():
            value = raw.get(header) if header else None
            canonical_row[canonical] = _normalize_cell(value)
        if not canonical_row.get("imei"):
            detected = _detect_imei(raw.values())
            if detected:
                canonical_row["imei"] = detected
        canonical_rows.append(canonical_row)
    return canonical_rows


def _detect_column_type(samples: list[str]) -> str | None:
    if not samples:
        return None
    if all(IMEI_PATTERN.fullmatch(sample) for sample in samples[:5]):
        return "imei"
    boolean_matches = sum(1 for sample in samples if _looks_boolean(sample))
    if boolean_matches and boolean_matches == len(samples):
        return "booleano"
    if boolean_matches and len(samples) >= 5 and boolean_matches / len(samples) >= 0.8:
        return "booleano"
    numeric = sum(1 for sample in samples if _looks_numeric(sample))
    if numeric == len(samples):
        return "numero"
    date_like = sum(1 for sample in samples if _looks_like_date(sample))
    if date_like >= len(samples) // 2 and date_like > 0:
        return "fecha"
    return "texto"


def _looks_numeric(value: str) -> bool:
    try:
        Decimal(value.replace(",", "."))
        return True
    except Exception:
        return False


def _looks_like_date(value: str) -> bool:
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            datetime.strptime(value, pattern)
            return True
        except ValueError:
            continue
    return False


def _normalize_boolean_token(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    normalized = "".join(
        char for char in normalized if not unicodedata.combining(char))
    return normalized.strip().lower()


def _looks_boolean(value: str) -> bool:
    token = _normalize_boolean_token(value)
    if not token:
        return False
    return token in BOOLEAN_TRUE_VALUES or token in BOOLEAN_FALSE_VALUES


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    normalized = "".join(
        char for char in normalized if not unicodedata.combining(char))
    normalized = normalized.lower()
    cleaned = re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")
    return cleaned


def _normalize_cell(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.strip()
        return normalized or None
    if isinstance(value, (int, float, Decimal)):
        return str(value)
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip() or None


def _normalize_optional(value: Any) -> str | None:
    normalized = _normalize_cell(value)
    return normalized


def _detect_imei(values: Iterable[Any]) -> str | None:
    for value in values:
        normalized = _normalize_cell(value)
        if normalized and IMEI_PATTERN.fullmatch(normalized):
            return normalized
    return None


def _resolve_estado_comercial(
    value: Any,
) -> tuple[models.CommercialState, str | None]:
    normalized = _normalize_cell(value)
    if not normalized:
        return models.CommercialState.NUEVO, None
    candidate_upper = normalized.upper()
    if candidate_upper in ESTADOS_COMERCIALES_VALIDOS:
        for candidate in (candidate_upper, candidate_upper.lower(), normalized):
            try:
                return models.CommercialState(candidate), None
            except ValueError:
                continue
    return models.CommercialState.NUEVO, normalized


def _parse_int(value: Any) -> int | None:
    text = _normalize_cell(value)
    if text is None:
        return None
    cleaned = text.replace(",", "").replace(" ", "")
    try:
        return int(Decimal(cleaned))
    except Exception:
        return None


def _parse_decimal(value: Any) -> Decimal | None:
    text = _normalize_cell(value)
    if text is None:
        return None
    cleaned = text.replace("$", "").replace(
        "€", "").replace(",", "").replace(" ", "")
    try:
        return Decimal(cleaned)
    except Exception:
        return None


def _parse_date(value: Any) -> date | None:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = _normalize_cell(value)
    if text is None:
        return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _validate_imei(value: Any) -> str | None:
    normalized = _normalize_cell(value)
    if normalized and IMEI_PATTERN.fullmatch(normalized):
        return normalized
    return None


def _generate_sku(store_code: str | None, modelo: str | None, row_index: int) -> str:
    prefix = store_code or "AUTO"
    modelo_part = _normalize_header(modelo or "modelo")[:6]
    return f"{prefix}-{modelo_part or 'EQ'}-{row_index:04d}"


def _generate_name(row: dict[str, Any]) -> str:
    parts = [row.get("marca"), row.get("modelo"),
             row.get("capacidad"), row.get("color")]
    return " ".join(filter(None, parts)) or "Producto importado"


def _is_row_incomplete(row: dict[str, Any]) -> bool:
    for field in CRITICAL_FIELDS:
        if not _normalize_cell(row.get(field)):
            return True
    for field in IMEI_IMPORTANT_FIELDS:
        if field in row and not _normalize_cell(row.get(field)):
            return True
    return False


def _row_has_data(values: Iterable[Any]) -> bool:
    for value in values:
        normalized = _normalize_cell(value)
        if normalized:
            return True
    return False
