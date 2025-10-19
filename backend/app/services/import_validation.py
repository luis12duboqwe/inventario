"""Validación avanzada de importaciones de inventario."""
from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from time import perf_counter
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from .. import models, schemas

REQUIRED_COLUMNS: tuple[str, ...] = ("tienda", "marca", "modelo", "sku", "cantidad")
NUMERIC_FIELDS: tuple[str, ...] = ("cantidad", "precio", "costo")
DATE_FIELDS: tuple[str, ...] = ("fecha_compra", "fecha_ingreso")


def validar_importacion(
    db: Session,
    *,
    registros: Sequence[dict[str, Any]],
    columnas_faltantes: Iterable[str],
    import_duration: float,
) -> schemas.ImportValidationSummary:
    """Analiza los registros importados y almacena incidencias detectadas."""

    start_time = perf_counter()
    now = datetime.utcnow()
    validations: list[models.ImportValidation] = []
    errores = 0
    advertencias = 0
    campos_faltantes = set(columnas_faltantes)

    for column in REQUIRED_COLUMNS:
        if column in columnas_faltantes:
            errores += 1
            validations.append(
                models.ImportValidation(
                    producto_id=None,
                    tipo="estructura",
                    severidad="error",
                    descripcion=f"Columna faltante: {column}",
                    fecha=now,
                    corregido=False,
                )
            )

    existing_imeis: dict[str, int] = {}
    imei_statement = select(models.Device.imei, models.Device.id).where(
        models.Device.imei.isnot(None)
    )
    for imei_value, device_id in db.execute(imei_statement):
        if imei_value:
            existing_imeis[str(imei_value)] = int(device_id)
    seen_imeis: dict[str, dict[str, Any]] = {}
    totals_por_tienda: defaultdict[int, int] = defaultdict(int)
    nombres_tiendas: dict[int, str] = {}

    for record in registros:
        row_index: int = record.get("row_index", 0)
        device_id: int | None = record.get("device_id")
        store_id: int | None = record.get("store_id")
        store_name: str | None = record.get("store_name")
        if store_id is not None and store_name:
            nombres_tiendas[store_id] = store_name
        imei = record.get("imei")
        serial = record.get("serial")
        for field in NUMERIC_FIELDS:
            raw_value = record.get(f"raw_{field}")
            parsed_value = record.get(field)
            if raw_value is not None and parsed_value is None:
                advertencias += 1
                validations.append(
                    models.ImportValidation(
                        producto_id=device_id,
                        tipo="estructura",
                        severidad="advertencia",
                        descripcion=f"Fila {row_index}: el campo '{field}' no tiene un formato numérico válido.",
                        fecha=now,
                        corregido=False,
                    )
                )
        for field in DATE_FIELDS:
            raw_date = record.get(field)
            if raw_date is None:
                continue
            if not isinstance(raw_date, date):
                advertencias += 1
                validations.append(
                    models.ImportValidation(
                        producto_id=device_id,
                        tipo="estructura",
                        severidad="advertencia",
                        descripcion=f"Fila {row_index}: la columna '{field}' no corresponde a una fecha válida.",
                        fecha=now,
                        corregido=False,
                    )
                )
        quantity = record.get("cantidad")
        if isinstance(quantity, int):
            totals_por_tienda[store_id or -1] += quantity
            if quantity < 0:
                errores += 1
                validations.append(
                    models.ImportValidation(
                        producto_id=device_id,
                        tipo="stock",
                        severidad="error",
                        descripcion=f"Fila {row_index}: el stock importado es negativo ({quantity}).",
                        fecha=now,
                        corregido=False,
                    )
                )
        if imei:
            if imei in seen_imeis:
                errores += 1
                first = seen_imeis[imei]["row_index"]
                validations.append(
                    models.ImportValidation(
                        producto_id=device_id,
                        tipo="duplicado",
                        severidad="error",
                        descripcion=f"Fila {row_index}: IMEI duplicado detectado (primera aparición en fila {first}).",
                        fecha=now,
                        corregido=False,
                    )
                )
            else:
                seen_imeis[imei] = {"row_index": row_index, "device_id": device_id}
            existing_id = existing_imeis.get(imei)
            if existing_id is not None and existing_id != device_id:
                errores += 1
                validations.append(
                    models.ImportValidation(
                        producto_id=device_id,
                        tipo="duplicado",
                        severidad="error",
                        descripcion=f"Fila {row_index}: el IMEI {imei} ya está registrado en el dispositivo {existing_id}.",
                        fecha=now,
                        corregido=False,
                    )
                )
        fecha_compra = record.get("fecha_compra")
        fecha_ingreso = record.get("fecha_ingreso")
        if isinstance(fecha_compra, date) and isinstance(fecha_ingreso, date):
            if fecha_compra > fecha_ingreso:
                advertencias += 1
                validations.append(
                    models.ImportValidation(
                        producto_id=device_id,
                        tipo="fechas",
                        severidad="advertencia",
                        descripcion=(
                            f"Fila {row_index}: la fecha de compra ({fecha_compra.isoformat()}) es posterior a la fecha de ingreso "
                            f"({fecha_ingreso.isoformat()})."
                        ),
                        fecha=now,
                        corregido=False,
                    )
                )
        if not imei and not serial:
            advertencias += 1
            validations.append(
                models.ImportValidation(
                    producto_id=device_id,
                    tipo="identificadores",
                    severidad="advertencia",
                    descripcion=f"Fila {row_index}: el registro carece de IMEI y número de serie.",
                    fecha=now,
                    corregido=False,
                )
            )

    if totals_por_tienda:
        store_ids = [store_id for store_id in totals_por_tienda.keys() if store_id >= 0]
        if store_ids:
            statement = (
                select(models.Device.store_id, func.sum(models.Device.quantity))
                .where(models.Device.store_id.in_(store_ids))
                .group_by(models.Device.store_id)
            )
            registros_actuales = {row.store_id: int(row[1] or 0) for row in db.execute(statement)}
            for store_id in store_ids:
                importado = totals_por_tienda.get(store_id, 0)
                registrado = registros_actuales.get(store_id, 0)
                if importado != registrado:
                    advertencias += 1
                    nombre = nombres_tiendas.get(store_id, f"Sucursal {store_id}")
                    diferencia = registrado - importado
                    validations.append(
                        models.ImportValidation(
                            producto_id=None,
                            tipo="desbalance_tiendas",
                            severidad="advertencia",
                            descripcion=(
                                f"{nombre}: desbalance entre el inventario importado ({importado}) y el registrado ({registrado}). "
                                f"Diferencia: {diferencia:+d} unidades."
                            ),
                            fecha=now,
                            corregido=False,
                        )
                    )

    if validations:
        db.add_all(validations)
        db.commit()

    validation_duration = perf_counter() - start_time
    return schemas.ImportValidationSummary(
        registros_revisados=len(registros),
        advertencias=advertencias,
        errores=errores,
        campos_faltantes=sorted(campos_faltantes),
        tiempo_total=round(import_duration + validation_duration, 2),
    )


def _sanitize_text(value: str) -> str:
    return value.replace("\n", " ").replace("\r", " ")


def export_validations_to_excel(
    validations: Sequence[models.ImportValidation],
    summary: schemas.ImportValidationSummary,
) -> BytesIO:
    """Genera un libro de Excel con las validaciones registradas."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Validaciones"
    headers = [
        "ID",
        "Producto",
        "Sucursal",
        "Tipo",
        "Severidad",
        "Descripción",
        "Fecha",
        "Corregido",
    ]
    sheet.append(headers)
    header_style = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_style
        cell.alignment = Alignment(horizontal="center")
    for validation in validations:
        device = validation.device
        product_label = "-"
        store_label = "-"
        if device is not None:
            product_label = f"{device.sku} · {device.name}"
            store_label = getattr(device.store, "name", store_label)
        sheet.append(
            [
                validation.id,
                product_label,
                store_label,
                validation.tipo,
                validation.severidad,
                _sanitize_text(validation.descripcion),
                validation.fecha.isoformat(),
                "Sí" if validation.corregido else "No",
            ]
        )
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max_length + 4, 60)

    summary_sheet = workbook.create_sheet("Resumen")
    summary_sheet.append(["Métrica", "Valor"])
    summary_sheet["A1"].font = Font(color="FFFFFF", bold=True)
    summary_sheet["B1"].font = Font(color="FFFFFF", bold=True)
    summary_sheet["A1"].fill = header_style
    summary_sheet["B1"].fill = header_style
    summary_sheet.append(["Registros revisados", summary.registros_revisados])
    summary_sheet.append(["Advertencias", summary.advertencias])
    summary_sheet.append(["Errores", summary.errores])
    summary_sheet.append(["Campos faltantes", ", ".join(summary.campos_faltantes) or "-"])
    summary_sheet.append(["Tiempo total (s)", summary.tiempo_total or 0])
    for column in summary_sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        summary_sheet.column_dimensions[column[0].column_letter].width = min(max_length + 4, 50)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def export_validations_to_pdf(
    validations: Sequence[models.ImportValidation],
    summary: schemas.ImportValidationSummary,
) -> BytesIO:
    """Genera un PDF con el detalle de las validaciones."""

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    heading_style = ParagraphStyle(
        "Heading",
        parent=styles["Heading1"],
        textColor=colors.HexColor("#38bdf8"),
        fontSize=16,
        spaceAfter=12,
    )
    body_style = styles["BodyText"]
    body_style.textColor = colors.HexColor("#1f2937")

    content = [Paragraph("Reporte de validaciones de importación", heading_style)]
    summary_table = Table(
        [
            ["Registros revisados", str(summary.registros_revisados)],
            ["Advertencias", str(summary.advertencias)],
            ["Errores", str(summary.errores)],
            ["Campos faltantes", ", ".join(summary.campos_faltantes) or "-"],
            ["Tiempo total (s)", str(summary.tiempo_total or 0)],
        ],
        colWidths=[180, 360],
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#e2e8f0")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#0f172a")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#38bdf8")),
                ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#38bdf8")),
            ]
        )
    )
    content.extend([summary_table, Spacer(1, 12)])

    if validations:
        table_data = [
            [
                "ID",
                "Producto",
                "Sucursal",
                "Tipo",
                "Severidad",
                "Descripción",
                "Fecha",
                "Corregido",
            ]
        ]
        for validation in validations:
            device = validation.device
            product_label = "-"
            store_label = "-"
            if device is not None:
                product_label = f"{device.sku} · {device.name}"
                store_label = getattr(device.store, "name", store_label)
            table_data.append(
                [
                    str(validation.id),
                    product_label,
                    store_label,
                    validation.tipo,
                    validation.severidad,
                    _sanitize_text(validation.descripcion),
                    validation.fecha.strftime("%d/%m/%Y %H:%M"),
                    "Sí" if validation.corregido else "No",
                ]
            )
        validations_table = Table(table_data, colWidths=[40, 120, 100, 70, 70, 200, 80, 60])
        validations_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
                    ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#0f172a")),
                    ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#38bdf8")),
                    ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#38bdf8")),
                ]
            )
        )
        content.append(validations_table)
    else:
        content.append(Paragraph("No se registran validaciones pendientes.", body_style))

    doc.build(content)
    buffer.seek(0)
    return buffer


def _normalize_numeric(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    try:
        normalized = str(value).replace(",", ".").strip()
        if not normalized:
            return None
        return Decimal(normalized)
    except Exception:
        return None


def build_record(
    *,
    row_index: int,
    store_id: int | None,
    store_name: str | None,
    imei: str | None,
    serial: str | None,
    raw_cantidad: Any,
    parsed_cantidad: int | None,
    raw_precio: Any,
    parsed_precio: Decimal | None,
    raw_costo: Any,
    parsed_costo: Decimal | None,
    fecha_compra: date | None,
    fecha_ingreso: date | None,
    device_id: int | None,
) -> dict[str, Any]:
    """Construye un registro serializable para la validación avanzada."""

    return {
        "row_index": row_index,
        "store_id": store_id,
        "store_name": store_name,
        "imei": imei,
        "serial": serial,
        "raw_cantidad": raw_cantidad,
        "cantidad": parsed_cantidad,
        "raw_precio": raw_precio,
        "precio": parsed_precio,
        "raw_costo": raw_costo,
        "costo": parsed_costo,
        "fecha_compra": fecha_compra,
        "fecha_ingreso": fecha_ingreso,
        "device_id": device_id,
    }
