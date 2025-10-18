from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .. import schemas


def _build_pdf_document(title: str) -> tuple[list, SimpleDocTemplate, BytesIO]:  # type: ignore[type-arg]
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, title=title)
    styles = getSampleStyleSheet()

    heading_style = ParagraphStyle(
        "HeadingSoftmobile",
        parent=styles["Heading1"],
        textColor=colors.HexColor("#38bdf8"),
        fontSize=18,
    )

    timestamp = datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")
    elements: list = [  # type: ignore[var-annotated]
        Paragraph("Softmobile 2025 — Directorio de usuarios", heading_style),
        Spacer(1, 12),
        Paragraph(f"Generado automáticamente el {timestamp}", styles["Normal"]),
        Spacer(1, 18),
    ]
    return elements, document, buffer


def _build_pdf_table(data: list[list[str]]) -> Table:
    table = Table(data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#e2e8f0")),
                ("LINEBEFORE", (0, 0), (0, -1), 1, colors.HexColor("#1e293b")),
                ("LINEAFTER", (-1, 0), (-1, -1), 1, colors.HexColor("#1e293b")),
                ("LINEABOVE", (0, 0), (-1, 0), 1.2, colors.HexColor("#38bdf8")),
                ("LINEBELOW", (0, -1), (-1, -1), 1.2, colors.HexColor("#38bdf8")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return table


def render_user_directory_pdf(report: schemas.UserDirectoryReport) -> bytes:
    elements, document, buffer = _build_pdf_document("Softmobile - Directorio de usuarios")
    styles = getSampleStyleSheet()

    filters: list[str] = []
    if report.filters.search:
        filters.append(f"Búsqueda: {report.filters.search}")
    if report.filters.role:
        filters.append(f"Rol: {report.filters.role}")
    if report.filters.status != "all":
        estado_map = {
            "active": "Activos",
            "inactive": "Inactivos",
            "locked": "Bloqueados",
        }
        estado = estado_map.get(report.filters.status, report.filters.status.title())
        filters.append(f"Estado: {estado}")
    if report.filters.store_id is not None:
        filters.append(f"Sucursal ID: {report.filters.store_id}")

    if filters:
        elements.append(Paragraph(" | ".join(filters), styles["Italic"]))
        elements.append(Spacer(1, 12))

    summary_table = _build_pdf_table(
        [
            ["Indicador", "Valor"],
            ["Usuarios totales", str(report.totals.total)],
            ["Activos", str(report.totals.active)],
            ["Inactivos", str(report.totals.inactive)],
            ["Bloqueados", str(report.totals.locked)],
        ]
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    detail_table: list[list[str]] = [
        [
            "Correo",
            "Nombre",
            "Rol primario",
            "Roles asignados",
            "Sucursal",
            "Estado",
            "Teléfono",
            "Último acceso",
        ]
    ]

    for item in report.items:
        last_login = "—"
        if item.last_login_at:
            last_login = item.last_login_at.strftime("%d/%m/%Y %H:%M")
        detail_table.append(
            [
                item.username,
                item.full_name or "—",
                item.rol,
                ", ".join(item.roles) if item.roles else "—",
                item.store_name or ("Sucursal " + str(item.store_id) if item.store_id else "—"),
                "Activo" if item.is_active else "Inactivo",
                item.telefono or "—",
                last_login,
            ]
        )

    elements.append(_build_pdf_table(detail_table))
    document.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def render_user_directory_xlsx(report: schemas.UserDirectoryReport) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Usuarios"

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="0f172a")
    data_fill = PatternFill("solid", fgColor="111827")
    accent_fill = PatternFill("solid", fgColor="1d4ed8")

    worksheet["A1"] = "Softmobile 2025 — Directorio de usuarios"
    worksheet["A1"].font = Font(color="38bdf8", bold=True, size=16)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    worksheet["A2"] = f"Generado: {datetime.utcnow().strftime('%d/%m/%Y %H:%M UTC')}"
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    summary_headers = ["Indicador", "Valor"]
    summary_rows = [
        ("Usuarios totales", report.totals.total),
        ("Activos", report.totals.active),
        ("Inactivos", report.totals.inactive),
        ("Bloqueados", report.totals.locked),
    ]

    worksheet.append(summary_headers)
    header_row = worksheet.max_row
    for column_index in range(1, len(summary_headers) + 1):
        cell = worksheet.cell(row=header_row, column=column_index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for label, value in summary_rows:
        worksheet.append([label, value])
        for column_index in range(1, 3):
            cell = worksheet.cell(row=worksheet.max_row, column=column_index)
            cell.fill = data_fill
            cell.font = Font(color="e2e8f0")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    worksheet.append([])

    headers = [
        "Correo",
        "Nombre",
        "Rol primario",
        "Roles asignados",
        "Sucursal",
        "Estado",
        "Teléfono",
        "Último acceso",
    ]
    worksheet.append(headers)
    for column_index in range(1, len(headers) + 1):
        cell = worksheet.cell(row=worksheet.max_row, column=column_index)
        cell.font = header_font
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in report.items:
        last_login = ""
        if item.last_login_at:
            last_login = item.last_login_at.strftime("%d/%m/%Y %H:%M")
        worksheet.append(
            [
                item.username,
                item.full_name or "",
                item.rol,
                ", ".join(item.roles) if item.roles else "",
                item.store_name or (f"Sucursal {item.store_id}" if item.store_id else ""),
                "Activo" if item.is_active else "Inactivo",
                item.telefono or "",
                last_login,
            ]
        )

    for column_index, _header in enumerate(headers, start=1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = 20

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


__all__ = [
    "render_user_directory_pdf",
    "render_user_directory_xlsx",
]
