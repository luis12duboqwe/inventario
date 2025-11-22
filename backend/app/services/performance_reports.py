from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .. import schemas


def _build_table(table_data: list[list[str]]) -> Table:
    table = Table(table_data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#e2e8f0")),
                ("LINEABOVE", (0, 0), (-1, 0), 1, colors.HexColor("#38bdf8")),
                ("LINEBELOW", (0, -1), (-1, -1), 1, colors.HexColor("#38bdf8")),
                ("ALIGN", (1, 1), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )
    return table


def _build_paragraphs(title: str):
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, title=title)
    styles = getSampleStyleSheet()
    subtitle_style = ParagraphStyle(
        "DarkSubtitle",
        parent=styles["Heading2"],
        textColor=colors.HexColor("#38bdf8"),
    )
    body_style = ParagraphStyle(
        "DarkBody",
        parent=styles["BodyText"],
        textColor=colors.HexColor("#e2e8f0"),
    )
    now_label = datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")
    elements: list = [  # type: ignore[var-annotated]
        Paragraph("Softmobile 2025 — Reporte consolidado", styles["Title"]),
        Spacer(1, 12),
        Paragraph(f"Generado automáticamente el {now_label}", body_style),
        Spacer(1, 16),
    ]
    return buffer, document, subtitle_style, body_style, elements


def render_financial_report_pdf(report: schemas.FinancialPerformanceReport) -> bytes:
    buffer, document, subtitle_style, body_style, elements = _build_paragraphs(
        "Softmobile - Reporte financiero"
    )
    totals = report.totals
    totals_table = _build_table(
        [
            ["Indicador", "Valor"],
            ["Ingresos", f"${totals.revenue:.2f}"],
            ["Costo", f"${totals.cost:.2f}"],
            ["Ganancia", f"${totals.profit:.2f}"],
            ["Margen %", f"{totals.margin_percent:.2f}%"],
        ]
    )
    elements.append(Paragraph("Totales consolidados", subtitle_style))
    elements.append(Spacer(1, 8))
    elements.append(totals_table)
    elements.append(Spacer(1, 14))

    elements.append(Paragraph("Márgenes por sucursal", subtitle_style))
    elements.append(Spacer(1, 6))
    profit_table: list[list[str]] = [
        ["Sucursal", "Ingresos", "Costo", "Ganancia", "Margen %"],
    ]
    for item in report.profit_by_store:
        profit_table.append(
            [
                item.store_name,
                f"${item.revenue:.2f}",
                f"${item.cost:.2f}",
                f"${item.profit:.2f}",
                f"{item.margin_percent:.2f}%",
            ]
        )
    elements.append(_build_table(profit_table))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Rotación de inventario", subtitle_style))
    elements.append(Spacer(1, 6))
    rotation_table: list[list[str]] = [
        ["Sucursal", "SKU", "Producto", "Rotación", "Vendidas", "Recibidas"],
    ]
    for item in report.rotation[:20]:
        rotation_table.append(
            [
                item.store_name,
                item.sku,
                item.name,
                f"{item.rotation_rate:.2f}",
                str(item.sold_units),
                str(item.received_units),
            ]
        )
    elements.append(_build_table(rotation_table))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Ventas por sucursal", subtitle_style))
    elements.append(Spacer(1, 6))
    sales_table: list[list[str]] = [["Sucursal", "Pedidos", "Unidades", "Ingresos"]]
    for item in report.sales_by_store:
        sales_table.append(
            [
                item.store_name,
                str(item.orders),
                str(item.units),
                f"${item.revenue:.2f}",
            ]
        )
    elements.append(_build_table(sales_table))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Ventas por categoría", subtitle_style))
    elements.append(Spacer(1, 6))
    category_table: list[list[str]] = [["Categoría", "Pedidos", "Unidades", "Ingresos"]]
    for item in report.sales_by_category:
        category_table.append(
            [
                item.category,
                str(item.orders),
                str(item.units),
                f"${item.revenue:.2f}",
            ]
        )
    elements.append(_build_table(category_table))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Tendencia por fecha", subtitle_style))
    elements.append(Spacer(1, 6))
    trend_table: list[list[str]] = [["Fecha", "Pedidos", "Unidades", "Ingresos"]]
    for item in report.sales_trend:
        trend_table.append(
            [
                item.date.isoformat(),
                str(item.orders),
                str(item.units),
                f"${item.revenue:.2f}",
            ]
        )
    elements.append(_build_table(trend_table))

    document.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def render_financial_report_xlsx(report: schemas.FinancialPerformanceReport) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen"
    bold = Font(bold=True)

    sheet.append(["Indicador", "Valor"])
    sheet["A1"].font = bold
    sheet["B1"].font = bold
    sheet.append(["Ingresos", report.totals.revenue])
    sheet.append(["Costo", report.totals.cost])
    sheet.append(["Ganancia", report.totals.profit])
    sheet.append(["Margen %", report.totals.margin_percent])

    sheet.append([])
    sheet.append(["Márgenes por sucursal"])
    sheet.append(["Sucursal", "Ingresos", "Costo", "Ganancia", "Margen %"])
    for cell in sheet[sheet.max_row]:
        cell.font = bold
    for item in report.profit_by_store:
        sheet.append(
            [
                item.store_name,
                item.revenue,
                item.cost,
                item.profit,
                item.margin_percent,
            ]
        )

    sheet.append([])
    sheet.append(["Rotación"])
    sheet.append(["Sucursal", "SKU", "Producto", "Rotación", "Vendidas", "Recibidas"])
    for cell in sheet[sheet.max_row]:
        cell.font = bold
    for item in report.rotation:
        sheet.append(
            [
                item.store_name,
                item.sku,
                item.name,
                item.rotation_rate,
                item.sold_units,
                item.received_units,
            ]
        )

    sheet.append([])
    sheet.append(["Ventas por sucursal"])
    sheet.append(["Sucursal", "Pedidos", "Unidades", "Ingresos"])
    for cell in sheet[sheet.max_row]:
        cell.font = bold
    for item in report.sales_by_store:
        sheet.append([item.store_name, item.orders, item.units, item.revenue])

    sheet.append([])
    sheet.append(["Ventas por categoría"])
    sheet.append(["Categoría", "Pedidos", "Unidades", "Ingresos"])
    for cell in sheet[sheet.max_row]:
        cell.font = bold
    for item in report.sales_by_category:
        sheet.append([item.category, item.orders, item.units, item.revenue])

    sheet.append([])
    sheet.append(["Tendencia por fecha"])
    sheet.append(["Fecha", "Pedidos", "Unidades", "Ingresos"])
    for cell in sheet[sheet.max_row]:
        cell.font = bold
    for item in report.sales_trend:
        sheet.append([item.date.isoformat(), item.orders, item.units, item.revenue])

    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def render_inventory_report_pdf(report: schemas.InventoryPerformanceReport) -> bytes:
    buffer, document, subtitle_style, body_style, elements = _build_paragraphs(
        "Softmobile - Reporte operativo de inventario"
    )
    elements.append(Paragraph("Rotación y ventas por sucursal/categoría", body_style))
    elements.append(Spacer(1, 14))

    elements.append(Paragraph("Rotación", subtitle_style))
    rotation_table: list[list[str]] = [
        ["Sucursal", "SKU", "Producto", "Rotación", "Vendidas", "Recibidas"],
    ]
    for item in report.rotation[:20]:
        rotation_table.append(
            [
                item.store_name,
                item.sku,
                item.name,
                f"{item.rotation_rate:.2f}",
                str(item.sold_units),
                str(item.received_units),
            ]
        )
    elements.append(_build_table(rotation_table))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Ventas por sucursal", subtitle_style))
    sales_table: list[list[str]] = [["Sucursal", "Pedidos", "Unidades", "Ingresos"]]
    for item in report.sales_by_store:
        sales_table.append(
            [
                item.store_name,
                str(item.orders),
                str(item.units),
                f"${item.revenue:.2f}",
            ]
        )
    elements.append(_build_table(sales_table))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Ventas por categoría", subtitle_style))
    category_table: list[list[str]] = [["Categoría", "Pedidos", "Unidades", "Ingresos"]]
    for item in report.sales_by_category:
        category_table.append(
            [
                item.category,
                str(item.orders),
                str(item.units),
                f"${item.revenue:.2f}",
            ]
        )
    elements.append(_build_table(category_table))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Margen por sucursal", subtitle_style))
    margin_table: list[list[str]] = [["Sucursal", "Ingresos", "Costo", "Margen %"]]
    for item in report.profit_by_store:
        margin_table.append(
            [
                item.store_name,
                f"${item.revenue:.2f}",
                f"${item.cost:.2f}",
                f"{item.margin_percent:.2f}%",
            ]
        )
    elements.append(_build_table(margin_table))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Tendencia por fecha", subtitle_style))
    trend_table: list[list[str]] = [["Fecha", "Pedidos", "Unidades", "Ingresos"]]
    for item in report.sales_trend:
        trend_table.append(
            [
                item.date.isoformat(),
                str(item.orders),
                str(item.units),
                f"${item.revenue:.2f}",
            ]
        )
    elements.append(_build_table(trend_table))

    document.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def render_inventory_report_xlsx(report: schemas.InventoryPerformanceReport) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventario"
    bold = Font(bold=True)

    sheet.append(["Rotación"])
    sheet.append(["Sucursal", "SKU", "Producto", "Rotación", "Vendidas", "Recibidas"])
    for cell in sheet[sheet.max_row]:
        cell.font = bold
    for item in report.rotation:
        sheet.append(
            [
                item.store_name,
                item.sku,
                item.name,
                item.rotation_rate,
                item.sold_units,
                item.received_units,
            ]
        )

    sheet.append([])
    sheet.append(["Ventas por sucursal"])
    sheet.append(["Sucursal", "Pedidos", "Unidades", "Ingresos"])
    for cell in sheet[sheet.max_row]:
        cell.font = bold
    for item in report.sales_by_store:
        sheet.append([item.store_name, item.orders, item.units, item.revenue])

    sheet.append([])
    sheet.append(["Ventas por categoría"])
    sheet.append(["Categoría", "Pedidos", "Unidades", "Ingresos"])
    for cell in sheet[sheet.max_row]:
        cell.font = bold
    for item in report.sales_by_category:
        sheet.append([item.category, item.orders, item.units, item.revenue])

    sheet.append([])
    sheet.append(["Margen por sucursal"])
    sheet.append(["Sucursal", "Ingresos", "Costo", "Ganancia", "Margen %"])
    for cell in sheet[sheet.max_row]:
        cell.font = bold
    for item in report.profit_by_store:
        sheet.append(
            [
                item.store_name,
                item.revenue,
                item.cost,
                item.profit,
                item.margin_percent,
            ]
        )

    sheet.append([])
    sheet.append(["Tendencia por fecha"])
    sheet.append(["Fecha", "Pedidos", "Unidades", "Ingresos"])
    for cell in sheet[sheet.max_row]:
        cell.font = bold
    for item in report.sales_trend:
        sheet.append([item.date.isoformat(), item.orders, item.units, item.revenue])

    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
