"""Generadores de libro fiscal mensual para ventas y compras."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from typing import Iterable, Sequence
from xml.etree.ElementTree import Element, SubElement, tostring

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .. import models, schemas

RATE_15 = Decimal("0.15")
RATE_18 = Decimal("0.18")
_TWO_PLACES = Decimal("0.01")
_FOUR_PLACES = Decimal("0.0001")


@dataclass(frozen=True)
class _CanonicalRecord:
    correlativo: int
    fecha: datetime
    documento: str
    contraparte: str | None
    detalle: str | None
    subtotal: Decimal
    impuesto: Decimal
    total: Decimal


def _quantize(value: Decimal) -> Decimal:
    return value.quantize(_TWO_PLACES, rounding=ROUND_HALF_UP)


def _format_currency(value: Decimal) -> str:
    normalized = float(_quantize(value))
    return f"L {normalized:,.2f}"


def _classify_tax(subtotal: Decimal, impuesto: Decimal) -> tuple[Decimal, Decimal, Decimal]:
    """Regresa tasas separadas para 15% y 18%, resto se marca como 0."""

    if subtotal <= Decimal("0") or impuesto <= Decimal("0"):
        return Decimal("0"), Decimal("0"), Decimal("0")

    ratio = (impuesto / subtotal).quantize(_FOUR_PLACES, rounding=ROUND_HALF_UP)
    diff_15 = abs(ratio - RATE_15)
    diff_18 = abs(ratio - RATE_18)

    if diff_15 <= Decimal("0.0015") and diff_15 <= diff_18:
        return RATE_15, Decimal("0"), Decimal("0")
    if diff_18 <= Decimal("0.0015") and diff_18 < diff_15:
        return Decimal("0"), RATE_18, Decimal("0")

    # Si no coincide exactamente asignamos a la tasa más cercana.
    if diff_15 <= diff_18:
        return RATE_15, Decimal("0"), Decimal("0")
    return Decimal("0"), RATE_18, Decimal("0")


def _build_entries(records: Sequence[_CanonicalRecord], filters: schemas.FiscalBookFilters) -> schemas.FiscalBookReport:
    entries: list[schemas.FiscalBookEntry] = []
    total_base_15 = Decimal("0")
    total_tax_15 = Decimal("0")
    total_base_18 = Decimal("0")
    total_tax_18 = Decimal("0")
    total_exempt = Decimal("0")
    total_general = Decimal("0")

    for record in records:
        base_15 = Decimal("0")
        tax_15 = Decimal("0")
        base_18 = Decimal("0")
        tax_18 = Decimal("0")
        base_exempt = Decimal("0")

        subtotal = _quantize(record.subtotal)
        impuesto = _quantize(record.impuesto)
        total = _quantize(record.total)

        if impuesto <= Decimal("0"):
            base_exempt = subtotal
        else:
            rate_15, rate_18, _ = _classify_tax(subtotal, impuesto)
            if rate_15:
                base_15 = subtotal
                tax_15 = impuesto
            elif rate_18:
                base_18 = subtotal
                tax_18 = impuesto
            else:
                base_exempt = subtotal

        entries.append(
            schemas.FiscalBookEntry(
                correlativo=record.correlativo,
                fecha=record.fecha,
                documento=record.documento,
                contraparte=record.contraparte,
                detalle=record.detalle,
                base_15=base_15,
                impuesto_15=tax_15,
                base_18=base_18,
                impuesto_18=tax_18,
                base_exenta=base_exempt,
                total=total,
            )
        )

        total_base_15 += base_15
        total_tax_15 += tax_15
        total_base_18 += base_18
        total_tax_18 += tax_18
        total_exempt += base_exempt
        total_general += total

    totals = schemas.FiscalBookTotals(
        registros=len(entries),
        base_15=_quantize(total_base_15),
        impuesto_15=_quantize(total_tax_15),
        total_15=_quantize(total_base_15 + total_tax_15),
        base_18=_quantize(total_base_18),
        impuesto_18=_quantize(total_tax_18),
        total_18=_quantize(total_base_18 + total_tax_18),
        base_exenta=_quantize(total_exempt),
        total_exento=_quantize(total_exempt),
        total_general=_quantize(total_general),
    )

    return schemas.FiscalBookReport(
        generated_at=datetime.utcnow(),
        filters=filters,
        totals=totals,
        entries=entries,
    )


def build_sales_fiscal_book(
    sales: Iterable[models.Sale],
    filters: schemas.FiscalBookFilters,
) -> schemas.FiscalBookReport:
    canonical: list[_CanonicalRecord] = []
    for idx, sale in enumerate(sorted(sales, key=lambda s: (s.created_at, s.id or 0)), start=1):
        store_name = sale.store.name if sale.store else f"Sucursal #{sale.store_id}"
        prefix = store_name.strip().upper()[:4].replace(" ", "") if store_name else "VENT"
        folio = f"{prefix or 'VENT'}-{(sale.id or idx):06d}"
        customer_name = sale.customer_name
        if not customer_name and sale.customer:
            customer_name = sale.customer.name
        canonical.append(
            _CanonicalRecord(
                correlativo=idx,
                fecha=sale.created_at,
                documento=folio,
                contraparte=customer_name,
                detalle=store_name,
                subtotal=Decimal(sale.subtotal_amount or Decimal("0")),
                impuesto=Decimal(sale.tax_amount or Decimal("0")),
                total=Decimal(sale.total_amount or Decimal("0")),
            )
        )
    return _build_entries(canonical, filters)


def build_purchases_fiscal_book(
    purchases: Iterable[models.Compra],
    filters: schemas.FiscalBookFilters,
) -> schemas.FiscalBookReport:
    canonical: list[_CanonicalRecord] = []
    for idx, purchase in enumerate(sorted(purchases, key=lambda p: (p.fecha, p.id_compra)), start=1):
        vendor_name = (
            purchase.proveedor.nombre
            if purchase.proveedor
            else f"Proveedor #{purchase.proveedor_id}"
        )
        folio = f"COMP-{purchase.id_compra:06d}"
        canonical.append(
            _CanonicalRecord(
                correlativo=idx,
                fecha=purchase.fecha,
                documento=folio,
                contraparte=vendor_name,
                detalle=purchase.forma_pago,
                subtotal=Decimal(purchase.total - purchase.impuesto),
                impuesto=Decimal(purchase.impuesto),
                total=Decimal(purchase.total),
            )
        )
    return _build_entries(canonical, filters)


def render_fiscal_book_pdf(report: schemas.FiscalBookReport) -> bytes:
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, title="Libro fiscal")
    styles = getSampleStyleSheet()
    heading_style = ParagraphStyle(
        "HeadingFiscal",
        parent=styles["Heading1"],
        textColor=colors.HexColor("#38bdf8"),
    )
    now_label = datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")

    elements: list = [  # type: ignore[var-annotated]
        Paragraph("Softmobile 2025 — Libro fiscal", heading_style),
        Spacer(1, 12),
        Paragraph(
            f"Generado el {now_label}. Tipo: {report.filters.book_type.upper()} ({report.filters.month:02d}/{report.filters.year})",
            styles["Normal"],
        ),
        Spacer(1, 18),
    ]

    summary_data = [
        ["Indicador", "Valor"],
        ["Registros", str(report.totals.registros)],
        ["Base 15%", _format_currency(report.totals.base_15)],
        ["ISV 15%", _format_currency(report.totals.impuesto_15)],
        ["Base 18%", _format_currency(report.totals.base_18)],
        ["ISV 18%", _format_currency(report.totals.impuesto_18)],
        ["Base exenta", _format_currency(report.totals.base_exenta)],
        ["Total general", _format_currency(report.totals.total_general)],
    ]
    summary_table = Table(summary_data, hAlign="LEFT")
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#e2e8f0")),
                ("LINEABOVE", (0, 0), (-1, 0), 1, colors.HexColor("#38bdf8")),
                ("LINEBELOW", (0, -1), (-1, -1), 1, colors.HexColor("#38bdf8")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ]
        )
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    table_data: list[list[str]] = [
        [
            "#",
            "Fecha",
            "Documento",
            "Contraparte",
            "Detalle",
            "Base 15%",
            "ISV 15%",
            "Base 18%",
            "ISV 18%",
            "Exento",
            "Total",
        ]
    ]

    for entry in report.entries:
        table_data.append(
            [
                str(entry.correlativo),
                entry.fecha.strftime("%d/%m/%Y"),
                entry.documento,
                entry.contraparte or "—",
                entry.detalle or "—",
                _format_currency(entry.base_15),
                _format_currency(entry.impuesto_15),
                _format_currency(entry.base_18),
                _format_currency(entry.impuesto_18),
                _format_currency(entry.base_exenta),
                _format_currency(entry.total),
            ]
        )

    details_table = Table(table_data, hAlign="LEFT")
    details_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#e2e8f0")),
                ("LINEABOVE", (0, 0), (-1, 0), 1, colors.HexColor("#38bdf8")),
                ("LINEBELOW", (0, -1), (-1, -1), 1, colors.HexColor("#38bdf8")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("ALIGN", (0, 1), (0, -1), "CENTER"),
                ("ALIGN", (5, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    elements.append(details_table)

    document.build(elements)
    buffer.seek(0)
    return buffer.read()


def render_fiscal_book_excel(report: schemas.FiscalBookReport) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Libro Fiscal"

    headers = [
        "#",
        "Fecha",
        "Documento",
        "Contraparte",
        "Detalle",
        "Base 15%",
        "ISV 15%",
        "Base 18%",
        "ISV 18%",
        "Exento",
        "Total",
    ]
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill("solid", fgColor="0f172a")
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, entry in enumerate(report.entries, start=2):
        sheet.cell(row=row_index, column=1, value=entry.correlativo)
        sheet.cell(row=row_index, column=2, value=entry.fecha.strftime("%Y-%m-%d"))
        sheet.cell(row=row_index, column=3, value=entry.documento)
        sheet.cell(row=row_index, column=4, value=entry.contraparte or "")
        sheet.cell(row=row_index, column=5, value=entry.detalle or "")
        sheet.cell(row=row_index, column=6, value=float(_quantize(entry.base_15)))
        sheet.cell(row=row_index, column=7, value=float(_quantize(entry.impuesto_15)))
        sheet.cell(row=row_index, column=8, value=float(_quantize(entry.base_18)))
        sheet.cell(row=row_index, column=9, value=float(_quantize(entry.impuesto_18)))
        sheet.cell(row=row_index, column=10, value=float(_quantize(entry.base_exenta)))
        sheet.cell(row=row_index, column=11, value=float(_quantize(entry.total)))

    total_row = len(report.entries) + 2
    sheet.cell(row=total_row, column=5, value="Totales").font = Font(bold=True)
    sheet.cell(row=total_row, column=6, value=float(report.totals.base_15))
    sheet.cell(row=total_row, column=7, value=float(report.totals.impuesto_15))
    sheet.cell(row=total_row, column=8, value=float(report.totals.base_18))
    sheet.cell(row=total_row, column=9, value=float(report.totals.impuesto_18))
    sheet.cell(row=total_row, column=10, value=float(report.totals.base_exenta))
    sheet.cell(row=total_row, column=11, value=float(report.totals.total_general))

    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 18

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def render_fiscal_book_xml(report: schemas.FiscalBookReport) -> BytesIO:
    root = Element(
        "LibroFiscal",
        tipo=report.filters.book_type,
        mes=f"{report.filters.month:02d}",
        anio=str(report.filters.year),
    )

    totals = SubElement(root, "Totales")
    SubElement(totals, "Registros").text = str(report.totals.registros)
    SubElement(totals, "Base15").text = f"{report.totals.base_15:.2f}"
    SubElement(totals, "ISV15").text = f"{report.totals.impuesto_15:.2f}"
    SubElement(totals, "Base18").text = f"{report.totals.base_18:.2f}"
    SubElement(totals, "ISV18").text = f"{report.totals.impuesto_18:.2f}"
    SubElement(totals, "BaseExenta").text = f"{report.totals.base_exenta:.2f}"
    SubElement(totals, "TotalGeneral").text = f"{report.totals.total_general:.2f}"

    registros = SubElement(root, "Registros")
    for entry in report.entries:
        registro = SubElement(registros, "Registro", correlativo=str(entry.correlativo))
        SubElement(registro, "Fecha").text = entry.fecha.strftime("%Y-%m-%d")
        SubElement(registro, "Documento").text = entry.documento
        SubElement(registro, "Contraparte").text = entry.contraparte or ""
        SubElement(registro, "Detalle").text = entry.detalle or ""
        SubElement(registro, "Base15").text = f"{entry.base_15:.2f}"
        SubElement(registro, "ISV15").text = f"{entry.impuesto_15:.2f}"
        SubElement(registro, "Base18").text = f"{entry.base_18:.2f}"
        SubElement(registro, "ISV18").text = f"{entry.impuesto_18:.2f}"
        SubElement(registro, "BaseExenta").text = f"{entry.base_exenta:.2f}"
        SubElement(registro, "Total").text = f"{entry.total:.2f}"

    payload = tostring(root, encoding="utf-8", xml_declaration=True)
    buffer = BytesIO(payload)
    buffer.seek(0)
    return buffer


__all__ = [
    "build_sales_fiscal_book",
    "build_purchases_fiscal_book",
    "render_fiscal_book_pdf",
    "render_fiscal_book_excel",
    "render_fiscal_book_xml",
]
