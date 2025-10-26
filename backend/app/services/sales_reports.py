"""Utilidades para generar reportes de ventas en PDF y Excel."""
from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from typing import Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .. import models, schemas


def _format_currency(value: Decimal | float | int) -> str:
    normalized = float(value)
    return f"${normalized:,.2f}"


def _build_pdf_table(data: list[list[str]]) -> Table:
    table = Table(data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#e2e8f0")),
                ("LINEABOVE", (0, 0), (-1, 0), 1, colors.HexColor("#38bdf8")),
                ("LINEBELOW", (0, -1), (-1, -1), 1, colors.HexColor("#38bdf8")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ]
        )
    )
    return table


def _build_pdf_document(title: str) -> tuple[list, SimpleDocTemplate, BytesIO]:  # type: ignore[type-arg]
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, title=title)
    styles = getSampleStyleSheet()
    now_label = datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")

    heading_style = ParagraphStyle(
        "Heading1Softmobile",
        parent=styles["Heading1"],
        textColor=colors.HexColor("#38bdf8"),
    )

    elements: list = [  # type: ignore[var-annotated]
        Paragraph("Softmobile 2025 — Reporte de ventas", heading_style),
        Spacer(1, 12),
        Paragraph(f"Generado automáticamente el {now_label}", styles["Normal"]),
        Spacer(1, 18),
    ]

    return elements, document, buffer


def _to_decimal(value: Decimal | float | int | None) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (ValueError, TypeError):
        return Decimal("0")


def build_sales_report(
    sales: list[models.Sale],
    filters: schemas.SalesReportFilters,
    *,
    audit_trails: Mapping[str, schemas.AuditTrailInfo] | None = None,
) -> schemas.SalesReport:
    subtotal_total = Decimal("0")
    tax_total = Decimal("0")
    total_amount = Decimal("0")
    cost_total = Decimal("0")
    daily_totals: dict[date, Decimal] = defaultdict(lambda: Decimal("0"))
    report_items: list[schemas.SalesReportItem] = []
    stores_map: dict[int | None, dict[str, object]] = {}
    users_map: dict[int | None, dict[str, object]] = {}
    products_map: dict[int, dict[str, object]] = {}
    first_sale_day: date | None = None
    last_sale_day: date | None = None

    for sale in sales:
        subtotal_total += sale.subtotal_amount
        tax_total += sale.tax_amount
        total_amount += sale.total_amount
        sale_day = sale.created_at.date()
        daily_totals[sale_day] += sale.total_amount
        if first_sale_day is None or sale_day < first_sale_day:
            first_sale_day = sale_day
        if last_sale_day is None or sale_day > last_sale_day:
            last_sale_day = sale_day

        customer_name = sale.customer_name or sale.customer.name if sale.customer else None
        performed_by = None
        if sale.performed_by:
            performed_by = sale.performed_by.full_name or sale.performed_by.username

        folio = f"VENTA-{sale.id:06d}"
        if sale.store and getattr(sale.store, "name", None):
            prefix = sale.store.name.strip().upper()[:4].replace(" ", "")
            if prefix:
                folio = f"{prefix}-{sale.id:06d}"

        item_models = [
            schemas.SaleItemResponse.model_validate(item, from_attributes=True)
            for item in sale.items
        ]

        last_action = audit_trails.get(str(sale.id)) if audit_trails else None

        report_items.append(
            schemas.SalesReportItem(
                sale_id=sale.id,
                folio=folio,
                store_name=sale.store.name if sale.store else f"Sucursal #{sale.store_id}",
                customer_name=customer_name,
                performed_by=performed_by,
                payment_method=sale.payment_method,
                subtotal=sale.subtotal_amount,
                tax=sale.tax_amount,
                total=sale.total_amount,
                created_at=sale.created_at,
                items=item_models,
                ultima_accion=last_action,
            )
        )

        store_entry = stores_map.setdefault(
            sale.store_id,
            {
                "name": sale.store.name if sale.store and sale.store.name else f"Sucursal #{sale.store_id}",
                "total": Decimal("0"),
                "count": 0,
            },
        )
        store_entry["total"] = _to_decimal(store_entry["total"]) + sale.total_amount
        store_entry["count"] = int(store_entry["count"]) + 1

        user_key = sale.performed_by_id
        user_name = performed_by or "—"
        user_entry = users_map.setdefault(
            user_key,
            {
                "name": user_name,
                "total": Decimal("0"),
                "count": 0,
            },
        )
        if not user_entry.get("name") and user_name:
            user_entry["name"] = user_name
        user_entry["total"] = _to_decimal(user_entry["total"]) + sale.total_amount
        user_entry["count"] = int(user_entry["count"]) + 1

        for item in sale.items:
            quantity_decimal = Decimal(item.quantity)
            raw_cost = getattr(item.device, "costo_unitario", None)
            unit_cost = _to_decimal(raw_cost)
            if raw_cost is None:
                unit_cost = _to_decimal(item.unit_price)
            line_cost = (unit_cost * quantity_decimal).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            cost_total += line_cost

            product_entry = products_map.setdefault(
                item.device_id,
                {
                    "sku": getattr(item.device, "sku", None),
                    "name": (
                        getattr(item.device, "name", None) or f"Producto #{item.device_id}"
                    ),
                    "units": 0,
                    "total": Decimal("0"),
                },
            )
            product_entry["units"] = int(product_entry["units"]) + item.quantity
            product_entry["total"] = _to_decimal(product_entry["total"]) + item.total_line

    daily_stats = [
        schemas.DashboardChartPoint(label=day.strftime("%d/%m/%Y"), value=float(amount))
        for day, amount in sorted(daily_totals.items())
    ]

    cost_total = cost_total.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    net_income = (total_amount - cost_total).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    start_day = filters.date_from.date() if filters.date_from else first_sale_day
    end_day = filters.date_to.date() if filters.date_to else last_sale_day
    daily_average = Decimal("0")
    if start_day and end_day:
        day_span = (end_day - start_day).days + 1
        if day_span > 0:
            daily_average = (total_amount / Decimal(day_span)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )

    totals = schemas.SalesReportTotals(
        count=len(sales),
        subtotal=subtotal_total,
        tax=tax_total,
        total=total_amount,
        cost=cost_total,
        net_income=net_income,
        daily_average=daily_average,
    )

    by_store = [
        schemas.SalesReportGroup(
            id=store_id,
            name=str(entry["name"]),
            total=_to_decimal(entry["total"]),
            count=int(entry["count"]),
        )
        for store_id, entry in sorted(
            stores_map.items(),
            key=lambda item: _to_decimal(item[1]["total"]),
            reverse=True,
        )
    ]

    by_user = [
        schemas.SalesReportGroup(
            id=user_id,
            name=str(entry["name"] or "—"),
            total=_to_decimal(entry["total"]),
            count=int(entry["count"]),
        )
        for user_id, entry in sorted(
            users_map.items(),
            key=lambda item: _to_decimal(item[1]["total"]),
            reverse=True,
        )
    ]

    top_products = [
        schemas.SalesReportProduct(
            product_id=product_id,
            sku=entry["sku"],
            name=str(entry["name"]),
            units=int(entry["units"]),
            total=_to_decimal(entry["total"]),
        )
        for product_id, entry in sorted(
            products_map.items(),
            key=lambda item: (int(item[1]["units"]), _to_decimal(item[1]["total"])),
            reverse=True,
        )[:10]
    ]

    return schemas.SalesReport(
        generated_at=datetime.utcnow(),
        filters=filters,
        totals=totals,
        daily_stats=daily_stats,
        items=report_items,
        by_store=by_store,
        by_user=by_user,
        top_products=top_products,
    )


def render_sales_report_pdf(report: schemas.SalesReport) -> bytes:
    elements, document, buffer = _build_pdf_document("Softmobile - Reporte de ventas")
    styles = getSampleStyleSheet()

    filters_parts: list[str] = []
    if report.filters.store_id:
        filters_parts.append(f"Sucursal #{report.filters.store_id}")
    if report.filters.customer_id:
        filters_parts.append(f"Cliente #{report.filters.customer_id}")
    if report.filters.performed_by_id:
        filters_parts.append(f"Usuario #{report.filters.performed_by_id}")
    if report.filters.product_id:
        filters_parts.append(f"Producto #{report.filters.product_id}")
    if report.filters.date_from or report.filters.date_to:
        start = report.filters.date_from.strftime("%d/%m/%Y") if report.filters.date_from else "∞"
        end = report.filters.date_to.strftime("%d/%m/%Y") if report.filters.date_to else "∞"
        filters_parts.append(f"Periodo: {start} → {end}")
    if report.filters.query:
        filters_parts.append(f"Búsqueda: {report.filters.query}")

    filters_text = ", ".join(filters_parts) if filters_parts else "Sin filtros aplicados"
    elements.append(Paragraph(filters_text, styles["Italic"]))
    elements.append(Spacer(1, 12))

    summary_table = _build_pdf_table(
        [
            ["Indicador", "Valor"],
            ["Operaciones", str(report.totals.count)],
            ["Subtotal", _format_currency(report.totals.subtotal)],
            ["Impuestos", _format_currency(report.totals.tax)],
            ["Total", _format_currency(report.totals.total)],
            ["Costo estimado", _format_currency(report.totals.cost)],
            ["Ingreso neto", _format_currency(report.totals.net_income)],
            ["Promedio diario", _format_currency(report.totals.daily_average)],
        ]
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    detail_table: list[list[str]] = [
        [
            "Folio",
            "Fecha",
            "Sucursal",
            "Cliente",
            "Total",
            "Impuesto",
            "Método",
            "Usuario",
            "Última acción",
        ]
    ]

    for item in report.items:
        last_action_label = "—"
        if item.ultima_accion:
            timestamp = item.ultima_accion.timestamp.strftime("%d/%m/%Y %H:%M")
            actor = item.ultima_accion.usuario or "—"
            last_action_label = f"{item.ultima_accion.accion} · {actor} · {timestamp}"
        detail_table.append(
            [
                item.folio,
                item.created_at.strftime("%d/%m/%Y %H:%M"),
                item.store_name,
                item.customer_name or "Mostrador",
                _format_currency(item.total),
                _format_currency(item.tax),
                item.payment_method.value,
                item.performed_by or "—",
                last_action_label,
            ]
        )

    elements.append(_build_pdf_table(detail_table))

    if report.by_store:
        elements.append(Spacer(1, 18))
        elements.append(Paragraph("Ventas por sucursal", styles["Heading3"]))
        store_table = _build_pdf_table(
            [["Sucursal", "Operaciones", "Total"]]
            + [
                [entry.name, str(entry.count), _format_currency(entry.total)]
                for entry in report.by_store
            ]
        )
        elements.append(store_table)

    if report.by_user:
        elements.append(Spacer(1, 18))
        elements.append(Paragraph("Ventas por usuario", styles["Heading3"]))
        user_table = _build_pdf_table(
            [["Usuario", "Operaciones", "Total"]]
            + [
                [entry.name, str(entry.count), _format_currency(entry.total)]
                for entry in report.by_user
            ]
        )
        elements.append(user_table)

    if report.top_products:
        elements.append(Spacer(1, 18))
        elements.append(Paragraph("Productos más vendidos", styles["Heading3"]))
        products_table = _build_pdf_table(
            [["Producto", "SKU", "Unidades", "Total"]]
            + [
                [
                    item.name,
                    item.sku or "—",
                    str(item.units),
                    _format_currency(item.total),
                ]
                for item in report.top_products
            ]
        )
        elements.append(products_table)

    if report.daily_stats:
        elements.append(Spacer(1, 18))
        elements.append(Paragraph("Ventas por día", styles["Heading3"]))
        daily_table = _build_pdf_table(
            [["Día", "Total"]]
            + [[entry.label, _format_currency(entry.value)] for entry in report.daily_stats]
        )
        elements.append(daily_table)

    document.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def render_sales_report_excel(report: schemas.SalesReport) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ventas"

    header_fill = PatternFill("solid", fgColor="0f172a")
    header_font = Font(color="FFFFFF", bold=True)

    sheet.append(["Resumen", "Valor"])
    sheet.cell(row=1, column=1).fill = header_fill
    sheet.cell(row=1, column=2).fill = header_fill
    sheet.cell(row=1, column=1).font = header_font
    sheet.cell(row=1, column=2).font = header_font

    summary_rows = [
        ("Operaciones", report.totals.count),
        ("Subtotal", _format_currency(report.totals.subtotal)),
        ("Impuestos", _format_currency(report.totals.tax)),
        ("Total", _format_currency(report.totals.total)),
        ("Costo estimado", _format_currency(report.totals.cost)),
        ("Ingreso neto", _format_currency(report.totals.net_income)),
        ("Promedio diario", _format_currency(report.totals.daily_average)),
    ]
    for label, value in summary_rows:
        sheet.append([label, value])

    sheet.append([])
    detail_headers = [
        "Folio",
        "Fecha",
        "Sucursal",
        "Cliente",
        "Total",
        "Impuesto",
        "Método",
        "Usuario",
        "Última acción",
    ]
    header_row = sheet.max_row + 1
    sheet.append(detail_headers)
    for col_idx, _ in enumerate(detail_headers, start=1):
        cell = sheet.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    for item in report.items:
        last_action_label = "—"
        if item.ultima_accion:
            timestamp = item.ultima_accion.timestamp.strftime("%d/%m/%Y %H:%M")
            actor = item.ultima_accion.usuario or "—"
            last_action_label = f"{item.ultima_accion.accion} · {actor} · {timestamp}"
        sheet.append(
            [
                item.folio,
                item.created_at.strftime("%d/%m/%Y %H:%M"),
                item.store_name,
                item.customer_name or "Mostrador",
                _format_currency(item.total),
                _format_currency(item.tax),
                item.payment_method.value,
                item.performed_by or "—",
                last_action_label,
            ]
        )

    for column in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(column)
        sheet.column_dimensions[column_letter].width = 18

    breakdown_sheet = workbook.create_sheet("Desgloses")
    breakdown_sheet.append(["Sucursal", "Operaciones", "Total"])
    for cell in breakdown_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for entry in report.by_store:
        breakdown_sheet.append(
            [entry.name, entry.count, _format_currency(entry.total)]
        )
    breakdown_sheet.append([])
    breakdown_sheet.append(["Usuario", "Operaciones", "Total"])
    header_row = breakdown_sheet.max_row
    for col_idx in range(1, 4):
        cell = breakdown_sheet.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
    for entry in report.by_user:
        breakdown_sheet.append(
            [entry.name, entry.count, _format_currency(entry.total)]
        )
    for column in range(1, breakdown_sheet.max_column + 1):
        breakdown_sheet.column_dimensions[get_column_letter(column)].width = 22

    products_sheet = workbook.create_sheet("Productos")
    products_sheet.append(["Producto", "SKU", "Unidades", "Total"])
    for cell in products_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for item in report.top_products:
        products_sheet.append(
            [item.name, item.sku or "—", item.units, _format_currency(item.total)]
        )
    for column in range(1, products_sheet.max_column + 1):
        width = 30 if column == 1 else 20
        products_sheet.column_dimensions[get_column_letter(column)].width = width

    daily_sheet = workbook.create_sheet("Diario")
    daily_sheet.append(["Día", "Total"])
    for cell in daily_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for entry in report.daily_stats:
        daily_sheet.append([entry.label, _format_currency(entry.value)])
    for column in range(1, daily_sheet.max_column + 1):
        column_letter = get_column_letter(column)
        daily_sheet.column_dimensions[column_letter].width = 20

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
