"""Herramientas para generar reportes PDF/Excel del módulo de compras."""
from collections import defaultdict
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .. import models, schemas
from .locale_helpers import format_dual_currency


def _format_currency(value: Decimal | float | int) -> str:
    return format_dual_currency(value)


def _build_pdf_table(data: list[list[str]]) -> Table:
    table = Table(data, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#e2e8f0")),
                ("LINEABOVE", (0, 0), (-1, 0), 1, colors.HexColor("#38bdf8")),
                ("LINEBELOW", (0, -1), (-1, -1), 1, colors.HexColor("#38bdf8")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ]
        )
    )
    return table


def _build_pdf_document(title: str) -> tuple[list, SimpleDocTemplate, BytesIO]:  # type: ignore[type-arg]
    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4, title=title)
    styles = getSampleStyleSheet()
    now_label = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")

    heading_style = ParagraphStyle(
        "Heading1SoftmobileCompras",
        parent=styles["Heading1"],
        textColor=colors.HexColor("#38bdf8"),
    )

    elements: list = [  # type: ignore[var-annotated]
        Paragraph("Softmobile 2025 — Reporte de compras", heading_style),
        Spacer(1, 12),
        Paragraph(f"Generado automáticamente el {now_label}", styles["Normal"]),
        Spacer(1, 18),
    ]

    return elements, document, buffer


def build_purchase_report(
    purchases: list[models.Compra],
    filters: schemas.PurchaseReportFilters,
) -> schemas.PurchaseReport:
    subtotal_total = Decimal("0")
    tax_total = Decimal("0")
    total_amount = Decimal("0")
    daily_totals: defaultdict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    report_items: list[schemas.PurchaseReportItem] = []

    for purchase in purchases:
        subtotal_value = purchase.total - purchase.impuesto
        subtotal_total += subtotal_value
        tax_total += purchase.impuesto
        total_amount += purchase.total
        day_label = purchase.fecha.strftime("%d/%m/%Y")
        daily_totals[day_label] += purchase.total

        vendor_name = (
            purchase.proveedor.nombre if purchase.proveedor else f"Proveedor #{purchase.proveedor_id}"
        )
        user_name = None
        if purchase.usuario:
            user_name = purchase.usuario.full_name or purchase.usuario.username

        folio = f"COMP-{purchase.id_compra:06d}"

        items = []
        for detail in purchase.detalles:
            product_name = detail.producto.name if detail.producto else None
            items.append(
                schemas.PurchaseRecordItemResponse(
                    id_detalle=detail.id_detalle,
                    producto_id=detail.producto_id,
                    cantidad=detail.cantidad,
                    costo_unitario=detail.costo_unitario,
                    subtotal=detail.subtotal,
                    producto_nombre=product_name,
                )
            )

        report_items.append(
            schemas.PurchaseReportItem(
                compra_id=purchase.id_compra,
                folio=folio,
                proveedor_nombre=vendor_name,
                usuario_nombre=user_name,
                forma_pago=purchase.forma_pago,
                estado=purchase.estado,
                subtotal=subtotal_value,
                impuesto=purchase.impuesto,
                total=purchase.total,
                fecha=purchase.fecha,
                items=items,
            )
        )

    daily_stats = [
        schemas.DashboardChartPoint(label=label, value=float(amount))
        for label, amount in sorted(daily_totals.items())
    ]

    totals = schemas.PurchaseReportTotals(
        count=len(purchases),
        subtotal=subtotal_total,
        impuesto=tax_total,
        total=total_amount,
    )

    return schemas.PurchaseReport(
        generated_at=datetime.now(timezone.utc),
        filters=filters,
        totals=totals,
        daily_stats=daily_stats,
        items=report_items,
    )


def render_purchase_report_pdf(report: schemas.PurchaseReport) -> bytes:
    elements, document, buffer = _build_pdf_document("Softmobile - Reporte de compras")
    styles = getSampleStyleSheet()

    filters_parts: list[str] = []
    if report.filters.proveedor_id:
        filters_parts.append(f"Proveedor #{report.filters.proveedor_id}")
    if report.filters.usuario_id:
        filters_parts.append(f"Usuario #{report.filters.usuario_id}")
    if report.filters.date_from or report.filters.date_to:
        start = report.filters.date_from.strftime("%d/%m/%Y") if report.filters.date_from else "∞"
        end = report.filters.date_to.strftime("%d/%m/%Y") if report.filters.date_to else "∞"
        filters_parts.append(f"Periodo: {start} → {end}")
    if report.filters.estado:
        filters_parts.append(f"Estado: {report.filters.estado}")
    if report.filters.query:
        filters_parts.append(f"Búsqueda: {report.filters.query}")

    filters_text = ", ".join(filters_parts) if filters_parts else "Sin filtros aplicados"
    elements.append(Paragraph(filters_text, styles["Italic"]))
    elements.append(Spacer(1, 12))

    summary_table = _build_pdf_table(
        [
            ["Indicador", "Valor"],
            ["Compras", str(report.totals.count)],
            ["Subtotal", _format_currency(report.totals.subtotal)],
            ["Impuestos", _format_currency(report.totals.impuesto)],
            ["Total", _format_currency(report.totals.total)],
        ]
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    detail_table: list[list[str]] = [
        [
            "Folio",
            "Fecha",
            "Proveedor",
            "Total",
            "Impuesto",
            "Pago",
            "Estado",
            "Usuario",
        ]
    ]

    for item in report.items:
        detail_table.append(
            [
                item.folio,
                item.fecha.strftime("%d/%m/%Y %H:%M"),
                item.proveedor_nombre,
                _format_currency(item.total),
                _format_currency(item.impuesto),
                item.forma_pago,
                item.estado,
                item.usuario_nombre or "—",
            ]
        )

    elements.append(_build_pdf_table(detail_table))

    if report.daily_stats:
        elements.append(Spacer(1, 18))
        daily_table = _build_pdf_table(
            [["Día", "Total"]]
            + [[entry.label, _format_currency(entry.value)] for entry in report.daily_stats]
        )
        elements.append(daily_table)

    document.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def render_purchase_report_excel(report: schemas.PurchaseReport) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Compras"

    header_fill = PatternFill("solid", fgColor="0f172a")
    header_font = Font(color="FFFFFF", bold=True)

    sheet.append(["Resumen", "Valor"])
    sheet.cell(row=1, column=1).fill = header_fill
    sheet.cell(row=1, column=2).fill = header_fill
    sheet.cell(row=1, column=1).font = header_font
    sheet.cell(row=1, column=2).font = header_font

    summary_rows = [
        ("Compras", report.totals.count),
        ("Subtotal", _format_currency(report.totals.subtotal)),
        ("Impuestos", _format_currency(report.totals.impuesto)),
        ("Total", _format_currency(report.totals.total)),
    ]
    for label, value in summary_rows:
        sheet.append([label, value])

    sheet.append([])
    detail_headers = [
        "Folio",
        "Fecha",
        "Proveedor",
        "Total",
        "Impuesto",
        "Pago",
        "Estado",
        "Usuario",
    ]
    header_row = sheet.max_row + 1
    sheet.append(detail_headers)
    for col_idx, _ in enumerate(detail_headers, start=1):
        cell = sheet.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    for item in report.items:
        sheet.append(
            [
                item.folio,
                item.fecha.strftime("%d/%m/%Y %H:%M"),
                item.proveedor_nombre,
                _format_currency(item.total),
                _format_currency(item.impuesto),
                item.forma_pago,
                item.estado,
                item.usuario_nombre or "—",
            ]
        )

    for column in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(column)
        sheet.column_dimensions[column_letter].width = 18

    daily_sheet = workbook.create_sheet("Diario")
    daily_sheet.append(["Día", "Total"])
    for cell in daily_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for entry in report.daily_stats:
        daily_sheet.append([entry.label, _format_currency(entry.value)])
    for column in range(1, daily_sheet.max_column + 1):
        column_letter = get_column_letter(column)
        daily_sheet.column_dimensions[column_letter].width = 20

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()
